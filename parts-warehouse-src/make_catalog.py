# -*- coding: utf-8 -*-
"""生成 category_catalog.xlsx (Excel 表一: 全部一级分类 + 子分类层级目录)。

用法: python make_catalog.py
输出: category_catalog.xlsx (项目根目录), 纯目录用途, 不存元器件。
"""

import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

from warehouse.config import CATEGORIES

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
OUT = os.path.join(BASE_DIR, "category_catalog.xlsx")

HEAD_FILL = PatternFill("solid", fgColor="2563EB")
HEAD_FONT = Font(color="FFFFFF", bold=True, size=12)
CAT_FONT = Font(bold=True, size=11)
ALT_FILL = PatternFill("solid", fgColor="F3F5F9")


def main():
    wb = Workbook()
    ws = wb.active
    ws.title = "分类总表"
    ws.append(["一级分类", "子分类", "说明"])
    for c in ws[1]:
        c.fill = HEAD_FILL
        c.font = HEAD_FONT
        c.alignment = Alignment(horizontal="center")

    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 44
    ws.column_dimensions["C"].width = 30

    row = 2
    for key, (name, subs) in CATEGORIES.items():
        if not subs:
            ws.append([name, "", "该一级分类无子分类"])
        else:
            ws.append([name, subs[0], ""])
        ws.cell(row=row, column=1).font = CAT_FONT
        if row % 2 == 0:
            for col in (1, 2, 3):
                ws.cell(row=row, column=col).fill = ALT_FILL
        row += 1
        for sub in subs[1:]:
            ws.append(["", sub, ""])
            if row % 2 == 0:
                for col in (1, 2, 3):
                    ws.cell(row=row, column=col).fill = ALT_FILL
            row += 1
        ws.append(["", "", ""])  # blank row between top-level categories
        row += 1

    ws.freeze_panes = "A2"
    wb.save(OUT)
    wb.close()
    n_subs = sum(len(s) for _, s in CATEGORIES.values())
    print(f"分类总表已生成: {OUT}")
    print(f"  一级分类 {len(CATEGORIES)} 个, 子分类 {n_subs} 个")


if __name__ == "__main__":
    main()
