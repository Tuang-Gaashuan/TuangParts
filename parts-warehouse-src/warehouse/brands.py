# -*- coding: utf-8 -*-
"""元器件仓库 — 品牌库聚合。

从全部子分类 Excel 的「品牌」列聚合出品牌视图:
  品牌名称 | 购买了多少种(条目数) | 购买了多少件(数量总和) | 该品牌经营的元器件类别

用途: 采购前快速回顾"哪些品牌我熟、买过什么元件", 优先向熟悉的品牌询价/下单。

同一品牌多种写法自动合并 (用户实测确认的口径):
  1. 英文+数字键相同:  FOJAN / FOJAN(富捷) / FOJAN富捷 / FOJAN(富健) → fojan
  2. 括号中文注释归并: 纯中文品牌与某英文品牌的括号注释一致 → 归入该组
     (如 "富捷" 的注释归属是 "FOJAN(富捷)" → 富捷 与 FOJAN 合并)
主名取出现次数最多的写法, 其余写法列为别名 (aliases)。

性能: 全量扫描结果按 (data 目录下所有 xlsx 的 mtime+size 指纹) 缓存,
数据未变化时打开品牌页/点品牌明细都是内存操作 (毫秒级)。
空品牌列 (未填写) 只计入 stats.no_brand, 绝不生成品牌条目。
"""

import os
import re
import time

from warehouse.config import CATEGORIES, primary_owner
from warehouse.excel_store import ExcelStore

# ── 缓存 (data_dir -> {"key": 指纹, "rows": 行索引, "data": 聚合结果}) ──
_cache = {}

# 缓存指纹中包含的文件类型
_XLSX_SUFFIX = ".xlsx"

# ── 品牌业务知识库 (品牌规范名 -> 主营业务描述) ──────────────
# 依据: 2026-08 上网核实 (立创商城/官网/百科) + 库存覆盖类别; 查无资料的标"(资料待核)"
BRAND_PROFILES = {
    "FOJAN(富捷)": "贴片电阻、电流采样电阻/分流器、排阻、贴片电容(MLCC)",
    "村田(Murata)": "片式电容(MLCC)、电感、磁珠、滤波器",
    "KOA": "贴片电阻、合金/采样电阻",
    "SAMSUNG(三星)": "片式电容(MLCC)",
    "厚声(UniOhm)": "贴片电阻、排阻(昆山厚声, Uniroyal/UniOhm)",
    "GOODWORK(固得沃克)": "二极管、MOS管",
    "FENGHUA": "片式电阻、片式电容(MLCC)(广东风华)",
    "YAGEO(国巨)": "贴片电阻、片式电容",
    "MDD(辰达半导体)": "整流/肖特基二极管、TVS保护器件",
    "R+O(宏嘉诚)": "二极管",
    "创基CBI": "二极管",
    "合科泰Hottech": "二极管、三极管、MOS管",
    "JIERR(捷而瑞)": "电流采样电阻/合金电阻",
    "台舟TECH PUBLIC": "二极管、MOS管",
    "JSMSEMI(杰盛微)": "二极管",
    "长江微电": "功率电感、NTC热敏电阻(东莞长江微电)",
    "YSUNK": "叠层高频电感(永胜电子)",
    "HXY MOSFET(华轩阳电子)": "MOSFET、电源管理芯片",
    "ST(意法半导体)": "单片机MCU、传感器、功率器件",
    "Torex": "LDO、DC-DC电源芯片(日本特瑞仕)",
    "SIkor(萨科微)": "二极管、MOS管",
    "BHFUSE(佰宏)": "自恢复/一次性保险丝",
    "乐山无线电LRC": "二极管、三极管",
    "东沃DOWO": "TVS/ESD保护二极管",
    "江苏长电(JCET)": "二极管、三极管、封装测试(长电科技)",
    "IDCHIP(英锐芯)": "运算放大器、接口芯片",
    "友台半导体UMW": "二极管、三极管",
    "静芯ElecSuper": "二极管、MOS管",
    "Econd(米朗电子)": "铝电解电容",
    "TKD(泰晶)": "石英晶振/谐振器(湖北泰晶科技)",
    "Coilank(驰兴)": "电感、磁珠(驰兴电感)",
    "华灿天禄": "IPEX/射频连接器",
    "TI(德州仪器)": "模拟芯片、电源管理、单片机MCU",
    "MaxLinear": "接口芯片(RS-485/收发器, 原Exar)",
    "瑞森半导体": "MOSFET功率器件",
    "国连": "FPC/FFC柔性连接器",
    "顾邦半导体GOSEMICON": "MOSFET功率器件",
    "SIT(芯力特)": "CAN/LIN收发器芯片(纳芯微旗下)",
    "XNRUSEMI(新锐)": "MOSFET功率器件",
    "捷茂微": "CAN接口芯片",
    "YXC扬兴科技": "石英晶振(扬兴科技)",
    "MagnTek(麦歌恩)": "磁性传感器(角度/电流/开关)",
    "WCH": "USB/以太网接口芯片、RISC-V单片机(沁恒微)",
    "Infineon": "功率半导体、MOSFET、传感器(英飞凌)",
    "MPS(芯源)": "电源管理芯片",
    "HI-LINK(海凌科)": "WiFi/蓝牙通信模块",
    # 查无权威资料的小众品牌: 业务按库存覆盖类别兜底, 不在此列
}


def _scan_key(data_dir: str) -> tuple:
    """数据目录指纹: 所有 xlsx 的 (相对路径, mtime, size)。

    任何文件增删改 (界面保存 / 用户直接改 Excel / 导入数据包) 都会改变指纹,
    用于缓存失效判断。只 stat 不读内容, 几百个文件毫秒级。
    """
    items = []
    for dirpath, _dirnames, filenames in os.walk(data_dir):
        for fn in filenames:
            if not fn.lower().endswith(_XLSX_SUFFIX):
                continue
            fp = os.path.join(dirpath, fn)
            try:
                st = os.stat(fp)
                items.append((os.path.relpath(fp, data_dir), st.st_mtime, st.st_size))
            except OSError:
                continue
    return tuple(sorted(items))


def _parse_qty(v) -> int:
    try:
        return int(float(str(v).strip()))
    except (ValueError, TypeError):
        return 0


def _norm_key(brand: str) -> str:
    """品牌归一化键: 英文+数字小写 (FOJAN(富捷) → fojan)。

    纯中文品牌无英文数字 → 返回空串 (由中文注释归并规则处理)。
    """
    return re.sub(r"[^a-z0-9]", "", str(brand or "").lower())


def _zh_comment(brand: str) -> str:
    """品牌名括号里的中文注释 (FOJAN(富捷) → 富捷; 意法半导体(ST) → 无)。"""
    m = re.search(r"[（(]([^（）()]*[\u4e00-\u9fff][^（）()]*)[）)]", str(brand or ""))
    return m.group(1) if m else ""


def _top(mapping: dict, n: int = 6) -> list:
    """按出现次数取前 n 个 key (类别标签按热度排序)。"""
    return [k for k, _ in sorted(mapping.items(), key=lambda kv: -kv[1])[:n]]


def _scan(data_dir: str) -> tuple:
    """全量扫描一次: 返回 (行索引, 聚合结果, 品牌写法→组键映射)。

    行索引 rows: [(subcat, owner_name, row)]  供品牌明细内存过滤。
    组键 bkey: {原始品牌写法: 组键}, 聚合与明细共用同一归并口径。
    """
    store = ExcelStore(data_dir)
    rows = []         # [(subcat, owner_name, row)]
    raw = {}          # 原始品牌写法 -> 聚合信息
    bkey = {}         # 品牌写法 -> 组键 (聚合/明细共用)
    no_brand = 0

    for subcat in store.all_subcat_counts():
        owner = primary_owner(subcat)
        owner_name = CATEGORIES[owner][0] if owner else ""
        _, data_rows = store.load(subcat)
        for r in data_rows:
            rows.append((subcat, owner_name, r))
            brand = (str(r[1]) if len(r) > 1 and r[1] is not None else "").strip()
            if not brand:
                no_brand += 1
                continue
            b = raw.setdefault(brand, {
                "brand": brand, "count": 0, "total_qty": 0,
                "owners": {}, "subcats": {}, "samples": [],
            })
            b["count"] += 1
            b["total_qty"] += _parse_qty(r[3] if len(r) > 3 else 0)
            if owner_name:
                b["owners"][owner_name] = b["owners"].get(owner_name, 0) + 1
            b["subcats"][subcat] = b["subcats"].get(subcat, 0) + 1
            name = str(r[0] if r and r[0] is not None else "").strip()
            if name and name not in b["samples"] and len(b["samples"]) < 8:
                b["samples"].append(name)
            bkey[brand] = _norm_key(brand) or brand

    # ── 中文注释归并: 纯中文品牌 → 括号注释匹配的英文品牌 ──
    # 先建"中文注释 -> [英文键品牌]"索引 (仅取有英文数字键的品牌)
    comment_idx = {}
    for brand in raw:
        if not _norm_key(brand):
            continue
        c = _zh_comment(brand)
        if c:
            comment_idx.setdefault(c, []).append(brand)
    # 纯中文品牌 (无英文数字键) 按注释归入对应品牌组
    for brand, b in list(raw.items()):
        if _norm_key(brand):
            continue
        targets = comment_idx.get(brand)
        if targets:
            best = max(targets, key=lambda t: raw[t]["count"])
            if best != brand:
                tgt = raw[best]
                tgt["count"] += b["count"]
                tgt["total_qty"] += b["total_qty"]
                for k, v in b["owners"].items():
                    tgt["owners"][k] = tgt["owners"].get(k, 0) + v
                for k, v in b["subcats"].items():
                    tgt["subcats"][k] = tgt["subcats"].get(k, 0) + v
                for s in b["samples"]:
                    if s not in tgt["samples"] and len(tgt["samples"]) < 8:
                        tgt["samples"].append(s)
                bkey[brand] = bkey[best]   # 明细查询也按同一组键归并
                del raw[brand]

    # ── 按组键分组 (纯中文品牌用原名做键) ──
    groups = {}
    for brand, b in raw.items():
        groups.setdefault(bkey[brand], []).append(b)

    brands = []
    for grp in groups.values():
        # 主名 = 出现次数最多; 同次数时含中文注释写法优先 (如 FOJAN(富捷) 优先于 FOJAN)
        grp.sort(key=lambda x: (-x["count"],
                                0 if _zh_comment(x["brand"]) else 1,
                                x["brand"]))
        main = grp[0]
        # 组内所有写法合并统计 (count/total_qty/类别/示例都要累加)
        merged = {
            "brand": main["brand"], "count": 0, "total_qty": 0,
            "owners": {}, "subcats": {}, "samples": [],
        }
        seen_samples = set()
        for b in grp:
            merged["count"] += b["count"]
            merged["total_qty"] += b["total_qty"]
            for k, v in b["owners"].items():
                merged["owners"][k] = merged["owners"].get(k, 0) + v
            for k, v in b["subcats"].items():
                merged["subcats"][k] = merged["subcats"].get(k, 0) + v
            for s in b["samples"]:
                if s not in seen_samples and len(merged["samples"]) < 8:
                    seen_samples.add(s)
                    merged["samples"].append(s)
        aliases = [b["brand"] for b in grp[1:]]
        # 业务: 知识库优先, 没有则按库存覆盖类别兜底
        business = BRAND_PROFILES.get(merged["brand"], "")
        if not business:
            owners_txt = "、".join(_top(merged["owners"], 6))
            business = f"{owners_txt}(按库存归集)" if owners_txt else "(资料待核)"
        brands.append({
            "brand": merged["brand"],
            "aliases": aliases,
            "count": merged["count"],
            "total_qty": merged["total_qty"],
            "owners": _top(merged["owners"], 6),
            "subcats": _top(merged["subcats"], 8),
            "samples": merged["samples"],
            "business": business,
        })
    brands.sort(key=lambda x: (-x["total_qty"], -x["count"]))

    data = {
        "brands": brands,
        "stats": {
            "brand_count": len(brands),
            "item_count": sum(b["count"] for b in brands),
            "total_qty": sum(b["total_qty"] for b in brands),
            "no_brand": no_brand,
        },
    }
    return rows, data, bkey


def _get_cached(data_dir: str) -> tuple:
    """按指纹返回缓存的 (rows, data, bkey); 缓存失效则重新扫描。"""
    key = _scan_key(data_dir)
    c = _cache.get(data_dir)
    if c is not None and c["key"] == key:
        return c["rows"], c["data"], c["bkey"]
    t0 = time.time()
    rows, data, bkey = _scan(data_dir)
    _cache[data_dir] = {"key": key, "rows": rows, "data": data, "bkey": bkey,
                        "scanned_at": time.time(), "scan_ms": int((time.time() - t0) * 1000)}
    return rows, data, bkey


def invalidate(data_dir: str):
    """显式失效缓存 (供写路径调用; 正常情况指纹检测已足够, 这是保险)。"""
    _cache.pop(data_dir, None)


def cache_info(data_dir: str) -> dict:
    """缓存状态 (调试用, 也供前端展示扫描耗时)。"""
    c = _cache.get(data_dir)
    if not c:
        return {"cached": False}
    return {"cached": True, "scanned_at": c["scanned_at"], "scan_ms": c["scan_ms"]}


def aggregate(data_dir: str) -> dict:
    """品牌聚合结果 (带缓存)。"""
    _, data, _bkey = _get_cached(data_dir)
    return data


def detail(data_dir: str, brand: str) -> dict:
    """某品牌(含所有别名写法)下的全部元器件行。

    返回 {items: [...], aliases: [同组其他写法...]} —— 从缓存行索引内存过滤。
    归并口径与聚合一致 (英文键 + 中文注释), 如查 FOJAN 能看到"富捷"行。
    """
    brand = (brand or "").strip()
    if not brand:
        return {"items": [], "aliases": []}
    rows, _data, bkey = _get_cached(data_dir)
    key = bkey.get(brand)
    if key is None:
        return {"items": [], "aliases": []}
    matches = []
    group_writings = {}   # 组内出现的原始写法 -> 行数 (算主名/别名用)
    for subcat, owner_name, r in rows:
        b = (str(r[1]) if len(r) > 1 and r[1] is not None else "").strip()
        if not b or bkey.get(b) != key:
            continue
        group_writings[b] = group_writings.get(b, 0) + 1
        matches.append({
            "subcat": subcat,
            "owner": owner_name,
            "row": len(matches),
            "name": str(r[0] or "") if len(r) > 0 else "",
            "brand": b,
            "package": str(r[2] or "") if len(r) > 2 else "",
            "qty": str(r[3] or "") if len(r) > 3 else "",
            "location": str(r[4] or "") if len(r) > 4 else "",
            "spec": str(r[6] or "") if len(r) > 6 else "",
            "datasheet": str(r[7] or "") if len(r) > 7 else "",
            "note": str(r[8] or "") if len(r) > 8 else "",
        })
    # 主名 = 组内行数最多的写法 (与 aggregate 的主名选择一致)
    main = max(group_writings, key=group_writings.get) if group_writings else brand
    aliases = [w for w in group_writings if w != main]
    return {"items": matches, "aliases": aliases}


def export_xlsx(data_dir: str) -> bytes:
    """品牌汇总导出为 xlsx (返回文件 bytes, 供下载)。"""
    from io import BytesIO
    from openpyxl import Workbook
    data = aggregate(data_dir)
    wb = Workbook()
    ws = wb.active
    ws.title = "品牌汇总"
    ws.append(["品牌", "别名", "种类数(条目)", "总数量(件)", "覆盖一级分类", "覆盖子分类", "型号示例"])
    for b in data["brands"]:
        ws.append([
            b["brand"], " / ".join(b["aliases"]), b["count"], b["total_qty"],
            " / ".join(b["owners"]), " / ".join(b["subcats"]),
            " / ".join(b["samples"]),
        ])
    for col, w in (("A", 26), ("B", 30), ("C", 12), ("D", 12), ("E", 40), ("F", 52), ("G", 70)):
        ws.column_dimensions[col].width = w
    buf = BytesIO()
    wb.save(buf)
    wb.close()
    return buf.getvalue()


def build_catalog(data_dir: str) -> str:
    """生成品牌库档案: data/品牌库.xlsx (品牌/业务/已购种类/总件数)。

    业务列优先保留用户已编辑过的内容 (按品牌名匹配旧文件),
    新品牌用知识库 BRAND_PROFILES, 再兜底覆盖类别。
    返回品牌库文件路径。
    """
    from openpyxl import Workbook, load_workbook
    data = aggregate(data_dir)
    # 读旧品牌库的业务列 (用户可能在 Excel 里改过, 保留)
    old_business = {}
    cat_path = os.path.join(data_dir, "品牌库.xlsx")
    if os.path.exists(cat_path):
        try:
            wb = load_workbook(cat_path, read_only=True, data_only=True)
            ws = wb.active
            for row in ws.iter_rows(min_row=2, values_only=True):
                if row and row[0]:
                    old_business[str(row[0]).strip()] = str(row[1] or "").strip()
            wb.close()
        except Exception:
            pass

    wb = Workbook()
    ws = wb.active
    ws.title = "品牌库"
    ws.append(["品牌", "业务", "已购种类", "总件数", "别名", "覆盖一级分类", "覆盖子分类", "型号示例"])
    for b in data["brands"]:
        biz = old_business.get(b["brand"], "") or b.get("business", "")
        ws.append([
            b["brand"], biz, b["count"], b["total_qty"],
            " / ".join(b["aliases"]), " / ".join(b["owners"]),
            " / ".join(b["subcats"]), " / ".join(b["samples"]),
        ])
    for col, w in (("A", 26), ("B", 46), ("C", 10), ("D", 10), ("E", 30),
                   ("F", 30), ("G", 44), ("H", 60)):
        ws.column_dimensions[col].width = w
    ws.freeze_panes = "A2"
    wb.save(cat_path)
    wb.close()
    return cat_path
