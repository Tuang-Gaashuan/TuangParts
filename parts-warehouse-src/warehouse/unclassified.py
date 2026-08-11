# -*- coding: utf-8 -*-
"""元器件仓库 — 未分类元件存储与手动归类。

未识别出分类的元件 (批量导入时 cat_key 匹配失败) 不再丢弃,
统一写入 data/未分类/未分类.xlsx (一个独立文件, 不占用分类体系),
用户可在「未分类」界面手动指定 一级分类+子分类, 移入正式仓库。

表头 = COMMON_FIELDS 标签 + 来源描述 (共 10 列)。
"""

import os

from openpyxl import Workbook, load_workbook

from warehouse.config import COMMON_FIELDS, safe_filename

UNCAT_DIR = "未分类"
UNCAT_FILE = "未分类.xlsx"

HEADERS = [label for _, label in COMMON_FIELDS] + ["来源描述"]


def uncat_path(data_dir: str) -> str:
    """未分类文件的绝对路径。"""
    return os.path.join(data_dir, UNCAT_DIR, UNCAT_FILE)


def load(data_dir: str) -> tuple[list, list]:
    """读取未分类数据, 返回 (表头, 行列表)。文件不存在返回 (HEADERS, [])。"""
    path = uncat_path(data_dir)
    if not os.path.exists(path):
        return list(HEADERS), []
    wb = load_workbook(path)
    ws = wb.active
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if any(c is not None and str(c).strip() != "" for c in row):
            rows.append(list(row))
    wb.close()
    return list(HEADERS), rows


def count(data_dir: str) -> int:
    """未分类元件条数。"""
    _, rows = load(data_dir)
    return len(rows)


def add(data_dir: str, items: list) -> int:
    """追加一批未分类元件。items: [{name, brand, package, qty, spec, raw}]。"""
    if not items:
        return 0
    headers, rows = load(data_dir)
    for it in items:
        rows.append([
            it.get("name", ""), it.get("brand", ""), it.get("package", ""),
            it.get("qty", "10"), "", "", it.get("spec", ""), "", "",
            it.get("raw", ""),
        ])
    _save(data_dir, headers, rows)
    return len(items)


def _save(data_dir: str, headers: list, rows: list):
    """写盘; rows 空时删除文件 (未分类为空则界面不显示)。"""
    path = uncat_path(data_dir)
    if not rows:
        if os.path.exists(path):
            os.remove(path)
            try:
                os.rmdir(os.path.dirname(path))
            except OSError:
                pass
        return
    os.makedirs(os.path.dirname(path), exist_ok=True)
    wb = Workbook()
    ws = wb.active
    ws.title = "未分类"
    ws.append(headers)
    for row in rows:
        padded = list(row) + [""] * (len(headers) - len(row))
        ws.append(padded[: len(headers)])
    wb.save(path)
    wb.close()


def remove(data_dir: str, indices: list) -> tuple[list, list]:
    """按行号(0 基)删除, 返回 (被删行列表, 剩余行列表)。"""
    headers, rows = load(data_dir)
    idx_set = set(int(i) for i in indices)
    removed, remaining = [], []
    for i, r in enumerate(rows):
        if i in idx_set:
            removed.append(r)
        else:
            remaining.append(r)
    _save(data_dir, headers, remaining)
    return removed, remaining


def assign(data_dir: str, indices: list, cat_key: str, subcat: str) -> int:
    """把未分类元件移入正式仓库: 从未分类删除 → 追加到目标子分类 Excel。

    返回移动条数。subcat 必须是非空且属于 cat_key (所有一级分类均有子分类)。
    """
    from warehouse.activity import record as record_activity
    from warehouse.excel_store import ExcelStore

    removed, remaining = remove(data_dir, indices)
    if not removed:
        return 0

    store = ExcelStore(data_dir)
    headers = [label for _, label in COMMON_FIELDS]
    old_headers, old_rows = store.load(subcat)
    new_rows = list(old_rows)
    for r in removed:
        # 未分类行: [名称,品牌,封装,数量,库位,子分类,规格,手册,备注,来源描述]
        new_rows.append([
            r[0] if len(r) > 0 else "",
            r[1] if len(r) > 1 else "",
            r[2] if len(r) > 2 else "",
            r[3] if len(r) > 3 else "",
            r[4] if len(r) > 4 else "",
            subcat,
            r[6] if len(r) > 6 else "",
            r[7] if len(r) > 7 else "",
            r[8] if len(r) > 8 else "",
        ])
    path = store.save(subcat, headers, new_rows)
    record_activity(data_dir, subcat, old_rows, new_rows, path=path)
    return len(removed)
