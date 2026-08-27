# -*- coding: utf-8 -*-
"""元器件仓库 — Excel 读写 (子分类文件模型).

结构:
  data/                      数据根目录
    <一级分类>/              按一级分类分子文件夹 (safe_filename 处理)
      <子分类>.xlsx          每个子分类一个 Excel 文件 (有数据才创建)

重名子分类 (如"磁珠"在 电感/EMI 下都有): 物理文件只建一份,
多个一级分类入口指向同一文件 (由 config.primary_owner 决定归属)。

空子分类不生成文件, 界面不显示。
"""

import os
from openpyxl import Workbook, load_workbook
from warehouse.config import CATEGORIES, fields_for, safe_filename, primary_owner

_STATS_CACHE = {}


class ExcelStore:
    """子分类 -> Excel 文件的读写层。"""

    def __init__(self, data_dir: str):
        self.data_dir = data_dir
        os.makedirs(data_dir, exist_ok=True)

    # ── 路径 ────────────────────────────────────────────
    def subcat_path(self, subcat: str) -> str:
        """子分类 Excel 的绝对路径 (归属第一个一级分类的文件夹)。"""
        owner = primary_owner(subcat)
        if owner is None:
            raise ValueError(f"子分类不存在: {subcat}")
        owner_dir = os.path.join(self.data_dir, safe_filename(CATEGORIES[owner][0]))
        return os.path.join(owner_dir, f"{safe_filename(subcat)}.xlsx")

    def owner_dir(self, key: str) -> str:
        return os.path.join(self.data_dir, safe_filename(CATEGORIES[key][0]))

    # ── 读 ──────────────────────────────────────────────
    def load(self, subcat: str) -> tuple[list, list]:
        """返回 (表头列表, 行数据列表)。文件不存在返回空。"""
        path = self.subcat_path(subcat)
        headers = [label for _, label in fields_for(subcat)]
        if not os.path.exists(path):
            return headers, []
        wb = load_workbook(path)
        ws = wb.active
        rows = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if any(c is not None and str(c).strip() != "" for c in row):
                rows.append(list(row))
        wb.close()
        return headers, rows

    # ── 写 (有数据才建文件) ─────────────────────────────
    def save(self, subcat: str, headers: list, rows: list) -> str:
        """保存子分类数据。rows 为空时删除文件 (空分类不显示)。返回文件路径。"""
        path = self.subcat_path(subcat)
        if not rows:
            # 空数据: 删除文件, 分类自动从界面消失
            if os.path.exists(path):
                os.remove(path)
                self._prune_empty_owner_dir(path)
            return path

        os.makedirs(os.path.dirname(path), exist_ok=True)
        wb = Workbook()
        ws = wb.active
        ws.title = safe_filename(subcat)[:31]
        ws.append(headers)
        for row in rows:
            padded = list(row) + [""] * (len(headers) - len(row))
            ws.append(padded[: len(headers)])
        for i, h in enumerate(headers, 1):
            ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = max(
                10, min(28, len(str(h)) * 2 + 4)
            )
        wb.save(path)
        wb.close()
        return path

    def _prune_empty_owner_dir(self, path: str):
        """删除子分类文件后, 若所属一级分类文件夹已空则删除该文件夹。"""
        d = os.path.dirname(path)
        try:
            if not os.listdir(d):
                os.rmdir(d)
        except OSError:
            pass

    # ── 统计 ────────────────────────────────────────────
    def count(self, subcat: str) -> int:
        _, rows = self.load(subcat)
        return len(rows)

    def subcats_with_data(self, key: str) -> list:
        """某一级分类下, 有数据的子分类列表 [(子分类名, 条数), ...]。"""
        result = []
        for subcat in CATEGORIES[key][1]:
            n = self.count(subcat)
            if n > 0:
                result.append((subcat, n))
        return result

    def category_total(self, key: str) -> int:
        """某一级分类下所有子分类的记录总数。"""
        return sum(n for _, n in self.subcats_with_data(key))

    def all_overview(self) -> dict:
        """总览: {一级分类key: 该分类记录总数}, 只含有数据的分类。"""
        counts = {subcat: len(rows) for subcat, rows in self._cached_stats_rows().items() if rows}
        result = {key: sum(counts.get(subcat, 0) for subcat in CATEGORIES[key][1]) for key in CATEGORIES}
        from warehouse import unclassified
        n = unclassified.count(self.data_dir)
        if n > 0:
            result["unclassified"] = n
        return result

    def _stats_fingerprint(self):
        files = []
        for root, _, names in os.walk(self.data_dir):
            for name in names:
                if name.lower().endswith(".xlsx"):
                    p = os.path.join(root, name)
                    try:
                        st = os.stat(p)
                        files.append((os.path.relpath(p, self.data_dir), st.st_mtime_ns, st.st_size))
                    except OSError:
                        pass
        return tuple(sorted(files))

    def _read_stats_rows(self):
        rows_by_subcat = {}
        for subcat in self._all_physical_subcats():
            path = self.subcat_path(subcat)
            if not os.path.exists(path):
                continue
            headers, rows = self.load(subcat)
            rows_by_subcat[subcat] = rows
        return rows_by_subcat

    def _all_physical_subcats(self):
        seen = set()
        for key in CATEGORIES:
            for subcat in CATEGORIES[key][1]:
                if subcat not in seen:
                    seen.add(subcat)
                    yield subcat

    def _cached_stats_rows(self):
        key = (self.data_dir, self._stats_fingerprint())
        cached = _STATS_CACHE.get(self.data_dir)
        if cached and cached[0] == key[1]:
            return cached[1]
        rows = self._read_stats_rows()
        _STATS_CACHE[self.data_dir] = (key[1], rows)
        return rows

    # ── 全局统计 ────────────────────────────────────────
    def all_subcat_counts(self) -> dict:
        """所有有数据的子分类 -> 条目数。"""
        result = {}
        for key in CATEGORIES:
            for subcat, n in self.subcats_with_data(key):
                # 重名子分类只统计一次 (物理文件唯一)
                result[subcat] = max(result.get(subcat, 0), n)
        return result

    def global_stats(self) -> dict:
        """浏览页统计，按 Excel 文件指纹缓存，避免每次打开首页重复读取全部库存。"""
        rows_by_subcat = self._cached_stats_rows()
        subcat_counts = {subcat: len(rows) for subcat, rows in rows_by_subcat.items() if rows}
        from warehouse import unclassified
        uncat_n = unclassified.count(self.data_dir)
        total_qty = 0
        for rows in rows_by_subcat.values():
            for r in rows:
                try:
                    total_qty += int(float(str(r[3]).strip()))
                except (ValueError, IndexError, TypeError):
                    pass
        cat_count = len({primary_owner(subcat) for subcat in subcat_counts}) + (1 if uncat_n else 0)
        top = max(subcat_counts.items(), key=lambda item: item[1], default=None)
        return {
            "categories": cat_count,
            "subcats": len(subcat_counts),
            "items": sum(subcat_counts.values()) + uncat_n,
            "total_qty": total_qty,
            "top_subcat": {"name": top[0], "count": top[1]} if top else None,
        }

    # 补货规则: 一级分类key -> 提醒阈值 (qty < 阈值 时提醒)
    #   None  = 该分类不参与补货提醒; 未列出的分类用默认阈值
    LOW_STOCK_RULES = {
        "tool": None,   # 工具类不做补货提醒
        "mcu": 2,       # 单片机/微控制器: 仅剩1个(<=1)时提醒, 即 qty < 2
    }

    def low_stock(self, threshold: int = 10) -> list:
        """库存不足扫描: 返回 [{subcat, name, qty, owner, threshold}]。

        按一级分类的补货规则计算 (工具类跳过, 单片机阈值2, 其余默认10)。
        """
        result = []
        rows_by_subcat = self._cached_stats_rows()
        for subcat, rows in rows_by_subcat.items():
            owner = primary_owner(subcat)
            if owner in self.LOW_STOCK_RULES and self.LOW_STOCK_RULES[owner] is None:
                continue  # 该分类明确不参与补货提醒 (如工具类)
            eff_threshold = self.LOW_STOCK_RULES.get(owner) or threshold
            for r in rows:
                if len(r) < 4:
                    continue
                try:
                    qty = int(float(str(r[3]).strip()))
                except (ValueError, TypeError):
                    qty = 0
                if qty < eff_threshold:
                    result.append({
                        "subcat": subcat,
                        "name": r[0] if r and r[0] else "(未命名)",
                        "qty": qty,
                        "owner": owner,
                        "threshold": eff_threshold,
                    })
        result.sort(key=lambda x: x["qty"])
        return result
