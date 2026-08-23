# -*- coding: utf-8 -*-
"""元器件仓库 — 批量导入 (Excel / txt / 粘贴文本)。

流程:
  1. 输入归一化: 粘贴文本 / txt 文件 → 文本行;  Excel 文件 → DataFrame
  2. LLM 批量解析: 每条记录 → 结构化字段 (型号/品牌/封装/数量/规格/子分类)
  3. 自动判断子分类: 从 585 个子分类中匹配最合适的
  4. 确认后批量写入各子分类 Excel

AI 接口配置与 ai_fill.py 相同 (data/ai_config.json)。
"""

import json
import os
import re
import difflib
from io import BytesIO
import zipfile

import xlrd
from openpyxl import load_workbook
from warehouse.config import CATEGORIES, COMMON_FIELDS, primary_owner
from warehouse.settings import load_settings, chat_completions_url

# 表头字段 key -> 中文标签 (与 config.COMMON_FIELDS 一致)
FIELD_LABELS = {k: v for k, v in COMMON_FIELDS}

# 位号/编号模式: 如 C1、R2、L3、U5、D1、LED1、SW2、FB1、XTAL1 (字母前缀 + 纯数字结尾)。
# 嘉立创等 EDA 导出的 BOM 常把 Comment 列填成位号编号, 不能当作元件名称。
PART_REF_RE = re.compile(
    r"^(C|R|L|D|U|Q|F|J|P|SW|LED|EC|FB|T|CN|XTAL|BZ|LS|BT|RN|JK|X|Y|Z|K)\d+$",
    re.IGNORECASE,
)


def _fix_part_ref_name(items: list) -> list:
    """名称是位号编号 (C1/R2/L3...) 时, 依次用 规格参数 → 原始描述 替换 (原地修改)。

    库存质量红线: 位号编号绝不允许作为元件名称入库。
    """
    for it in items:
        name = (it.get("name") or "").strip()
        if not name or not PART_REF_RE.match(name):
            continue
        spec = (it.get("spec") or "").strip()
        raw = (it.get("raw") or "").strip()
        if spec and not PART_REF_RE.match(spec):
            it["name"] = spec
        elif raw and not PART_REF_RE.match(raw):
            it["name"] = raw[:40]
    return items


# 不贴装标记: NC / N/C / DNP / 不贴(装) / 空贴。作为独立词出现才算
# (避免误伤真实型号, 如 NC7SZ08 安森美逻辑门)。
NC_TOKEN_RE = re.compile(
    r"(^|[/,\s(（])\s*(NC|DNP|N/C|不贴装|不贴|空贴)\s*([/,\s)）]|$)",
    re.IGNORECASE,
)


def _drop_nc_items(items: list) -> list:
    """剔除带 NC/DNP/不贴装 标记的行 (不贴装的元件不入库)。"""
    kept = []
    for it in items:
        text = " ".join(str(it.get(k) or "") for k in ("name", "spec", "raw"))
        if NC_TOKEN_RE.search(text):
            continue
        kept.append(it)
    return kept


def _normalize_quantity(value, default="10") -> str:
    """把 Excel 中常见的数量写法统一成可参与倍数计算的整数文本。"""
    text = str(value or "").strip()
    if not text:
        return default
    match = re.search(r"[-+]?\d+(?:\.\d+)?", text.replace(",", ""))
    if not match:
        return default
    try:
        number = float(match.group(0))
    except ValueError:
        return default
    if number < 0:
        return default
    return str(int(number)) if number.is_integer() else str(number)


# 电容容值模式: 数字 + 单位 (uF/nF/pF/mF/F), 大小写/µ 均可
CAP_VALUE_RE = re.compile(r"(\d+(?:\.\d+)?)\s*(u|µ|n|p|m)?F", re.IGNORECASE)


def _normalize_capacitor(items: list) -> list:
    """电容规格规范化: 名称只留容值, 混入名称的 耐压/材质 拆分到规格参数。

    如 "1UF@35V" → name="1uF", spec="35V";
       "10uF 50V X7R" → name="10uF", spec="50V X7R"。
    """
    for it in items:
        cat = it.get("cat_key") or ""
        subcat = it.get("subcat") or ""
        if cat != "capacitor" and "电容" not in str(subcat):
            continue  # 只处理电容类
        name = (it.get("name") or "").strip()
        m = CAP_VALUE_RE.search(name)
        if not m:
            continue
        cap = m.group(0).replace("µ", "u")
        cap = re.sub(r"([uUnNpPmM])F$", lambda x: x.group(1).lower() + "F", cap)  # 单位统一小写
        rest = (name[:m.start()] + " " + name[m.end():]).strip()
        rest = re.sub(r"[/@|,;（）()\s]+", " ", rest).strip()  # 分隔符归一
        it["name"] = cap   # 名称统一为规范化容值 (如 1uF)
        if rest:
            spec = (it.get("spec") or "").strip()
            if rest.lower() not in spec.lower():
                it["spec"] = (rest + " " + spec).strip() if spec else rest
    return items


def _merge_rows(old_rows: list, new_items: list, subcat: str) -> list:
    """合并重复行: (名称,品牌,封装,规格) 完全相同 → 数量累加, 备注拼接。

    old_rows 内部的历史重复也一并合并。返回合并后的行列表。
    列序 (COMMON_FIELDS): 名称0 品牌1 封装2 数量3 库位4 子分类5 规格6 手册7 备注8
    """
    def qty_of(row):
        try:
            return int(float(str(row[3]).strip()))
        except (TypeError, ValueError):
            return 0

    merged = {}
    order = []

    def add_row(row):
        key = (
            str(row[0] or "").strip(), str(row[1] or "").strip(),
            str(row[2] or "").strip(),
            str(row[6] if len(row) > 6 else "").strip(),
        )
        if key in merged:
            tgt = merged[key]
            tgt[3] = str(qty_of(tgt) + qty_of(row))
            if len(tgt) > 8 and len(row) > 8 and row[8]:
                tgt[8] = f"{tgt[8]} / {row[8]}" if tgt[8] else str(row[8])
        else:
            merged[key] = list(row)
            order.append(key)

    for r in old_rows:
        add_row(list(r))
    for it in new_items:
        add_row([
            it.get("name", ""), it.get("brand", ""), it.get("package", ""),
            it.get("qty", "10"), "", subcat, it.get("spec", ""), "", "",
        ])
    return [merged[k] for k in order]


def _category_tree_text() -> str:
    """生成分类树文本: '一级分类: 子分类1, 子分类2...' 供 LLM 参考。"""
    lines = []
    for key, (name, subs) in CATEGORIES.items():
        if subs:
            lines.append(f"{name}: {', '.join(subs)}")
        else:
            lines.append(f"{name}: (无子分类)")
    return "\n".join(lines)


def _match_subcat(name: str, subcat_list: list) -> str:
    """模糊匹配子分类名, 返回最接近的; 无匹配返回空。"""
    best, score = "", 0
    for s in subcat_list:
        r = difflib.SequenceMatcher(None, name.lower(), s.lower()).ratio()
        if r > score:
            best, score = s, r
    return best if score >= 0.5 else ""


def _match_cat_subcat(cat_name: str, subcat_name: str) -> tuple:
    """匹配一级分类+子分类, 返回 (cat_key, subcat)。匹配失败返回 (None, None)。"""
    # 先按一级分类名精确/模糊匹配
    cat_keys = [k for k, (n, _) in CATEGORIES.items() if n == cat_name]
    if not cat_keys:
        best, score = None, 0
        for k, (n, _) in CATEGORIES.items():
            r = difflib.SequenceMatcher(None, cat_name.lower(), n.lower()).ratio()
            if r > score:
                best, score = k, r
        cat_keys = [best] if best and score >= 0.4 else []
    if not cat_keys:
        return None, None
    key = cat_keys[0]
    subs = CATEGORIES[key][1]
    subcat = _match_subcat(subcat_name, subs) if subs else ""
    if not subs:
        subcat = ""   # 无子分类的大类
    return key, subcat


class BatchParser:
    """批量导入解析器。"""

    def __init__(self, api_key: str, base_url: str, model: str):
        import httpx
        self.httpx = httpx
        self.api_key = api_key
        self.base_url = base_url
        self.model = model
        self.dropped_nc = 0   # 最近一次解析剔除的 NC/不贴装 条数
        self.usage = {"prompt": 0, "completion": 0, "total": 0}   # tokens 累计

    # ── LLM 调用 ────────────────────────────────────────
    def _chat(self, messages: list) -> str:
        url = chat_completions_url(self.base_url)
        headers = {
            "Authorization": f"Bearer {self.api_key}",
            "Content-Type": "application/json",
        }
        payload = {"model": self.model, "messages": messages, "temperature": 0.05, "max_tokens": 8192}
        resp = self.httpx.post(url, headers=headers, json=payload, timeout=180)
        resp.raise_for_status()
        data = resp.json()
        # 累计 tokens 消耗
        usage = data.get("usage") or {}
        prompt_tk = usage.get("prompt_tokens") or 0
        comp_tk = usage.get("completion_tokens") or 0
        self.usage["prompt"] += prompt_tk
        self.usage["completion"] += comp_tk
        self.usage["total"] += usage.get("total_tokens") or (prompt_tk + comp_tk)
        return data["choices"][0]["message"]["content"]

    @staticmethod
    def _parse_json(text: str):
        text = text.strip()
        if text.startswith("```"):
            lines = text.split("\n")
            text = "\n".join(lines[1:-1] if lines[-1].strip() == "```" else lines[1:])
        m = re.search(r"\[.*\]|\{.*\}", text, re.DOTALL)
        if m:
            text = m.group(0)
        try:
            return json.loads(text)
        except json.JSONDecodeError:
            # 容错: LLM 输出被 max_tokens 截断时, 尝试补全残缺 JSON 数组
            # (从第一个 [ 截取到最后一个 "}" 或 "]", 补上结尾)
            start = text.find("[")
            if start >= 0:
                end = max(text.rfind("}"), text.rfind("]"))
                if end > start:
                    frag = text[start:end + 1]
                    if frag.endswith("}"):
                        frag += "]"
                    try:
                        return json.loads(frag)
                    except json.JSONDecodeError:
                        pass
            raise

    # ── 文本解析 (粘贴 / txt) ────────────────────────────
    def parse_text(self, text: str) -> list:
        """解析粘贴文本/txt 内容, 返回 [{name, brand, package, qty, spec, cat_key, subcat, raw}]。"""
        lines = [ln.strip() for ln in text.splitlines() if ln.strip()]
        if not lines:
            return []

        tree = _category_tree_text()
        items = []
        # 分批发送 (每批 40 行), 避免超长文本/输出截断导致只解析出前几条
        for start in range(0, len(lines), 40):
            batch = lines[start:start + 40]
            prompt = f"""你是电子元器件仓库管理员。下面有多条元器件描述, 每条可能是单独一行或编号条目。
请逐条解析, 输出 JSON 数组, 每项结构:
{{"index": 行号, "name": 型号或名称, "brand": 品牌, "package": 封装, "qty": 数量,
  "spec": 电气参数(阻值/容值/耐压/精度/频率等合并), "cat": 一级分类名, "subcat": 子分类名}}

分类体系 (一级分类: 子分类):
{tree}

规则:
- 每条输出一个对象, index 对应用户文本的行号(本批从0开始)。
- 解析不出品牌/封装/数量的字段用空字符串。
- qty 是数字或"个/片/只"等数量, 没有则默认 10。
- name 填型号或规格值(如 10uF、10kΩ、STM32F103C8T6); 若内容只是位号编号(如 C1、R2、L3、U5 这种 字母+数字), 不要当作 name。
- 电容: name 只填容值(如 10uF、100nF), 耐压(如 50V、25V)和材质(如 X7R、X5R、C0G、电解、钽)必须提取出来放进 spec, 格式如 "50V X7R"。
- cat/subcat 从上面分类体系里选最匹配的, 认不准时 cat 填最接近的, subcat 填空。
- 必须解析本批全部 {len(batch)} 条, 一条都不能漏。
- 只输出紧凑 JSON 数组(不要缩进换行), 不要多余文字。

用户文本:
{chr(10).join(f"[{i}] {ln}" for i, ln in enumerate(batch))}"""

            raw = self._chat([{"role": "user", "content": prompt}])
            data = self._parse_json(raw)
            if isinstance(data, dict):
                data = data.get("items", data.get("data", []))
            for d in data:
                if not isinstance(d, dict):
                    continue
                cat_key, subcat = _match_cat_subcat(
                    str(d.get("cat", "")), str(d.get("subcat", ""))
                )
                # raw: 原始行文本 (位号编号兜底用)
                raw_txt = ""
                try:
                    idx = int(d.get("index", -1))
                    if 0 <= idx < len(batch):
                        raw_txt = batch[idx]
                except (TypeError, ValueError):
                    pass
                items.append({
                    "name": str(d.get("name", "")).strip(),
                    "brand": str(d.get("brand", "")).strip(),
                    "package": str(d.get("package", "")).strip(),
                    "qty": _normalize_quantity(d.get("qty", "10")),
                    "spec": str(d.get("spec", "")).strip(),
                    "cat_key": cat_key,
                    "subcat": subcat,
                    "raw": raw_txt,
                })
        self.dropped_nc = len(items)
        items = _fix_part_ref_name(items)
        items = _normalize_capacitor(items)
        items = _drop_nc_items(items)
        self.dropped_nc -= len(items)
        return items

    # ── Excel 解析 ──────────────────────────────────────
    @staticmethod
    def _parse_structured_rows(header: list, data_rows: list) -> list | None:
        """直接解析具备明确字段的采购/库存表；表头不明确时返回 None 交给 AI。"""
        normalized = [str(cell).strip().lower().replace(" ", "") for cell in header]

        def column(*names):
            for index, value in enumerate(normalized):
                if any(name in value for name in names):
                    return index
            return None

        name_col = column("商品型号", "型号", "model", "partnumber", "mpn", "value", "值")
        brand_col = column("品牌", "manufacturer", "brand")
        package_col = column("封装格式", "封装", "package", "footprint")
        qty_col = column("订购数量", "数量", "quantity", "qty")
        desc_col = column("商品名称", "描述", "规格参数", "规格", "description", "comment", "remarks", "备注")
        type_col = column("商品类型", "元件类型", "type", "category")
        if name_col is None or (qty_col is None and desc_col is None):
            return None

        def value_at(row, index):
            if index is None or index >= len(row) or row[index] is None:
                return ""
            return str(row[index]).strip()

        items = []
        for row in data_rows:
            name = value_at(row, name_col)
            desc = value_at(row, desc_col)
            # 导出的采购报表常在明细后附“合计金额/说明”行；没有型号和描述即不是物料。
            if not name and not desc:
                continue
            raw = " ".join(value_at(row, index) for index in range(len(header))).strip()
            category_text = " ".join((value_at(row, type_col), desc, name))
            cat_key, subcat = None, ""
            for key, (_, subs) in CATEGORIES.items():
                for candidate in subs:
                    if candidate and candidate.lower() in category_text.lower():
                        cat_key, subcat = key, candidate
                        break
                if cat_key:
                    break
            if cat_key is None:
                from warehouse.rules import RuleParser
                classified = RuleParser().parse_line(category_text)
                if classified:
                    cat_key, subcat = classified["cat_key"], classified["subcat"]
            # 被动件优先用可检索的数值作名称；商品型号保留在规格，避免丢失精确料号。
            passive_value = ""
            if cat_key in {"capacitor", "resistor", "inductor"}:
                value_match = re.search(
                    r"(?<![A-Za-z0-9])(\d+(?:\.\d+)?\s*(?:[pnumkKMµ]?\s*(?:F|H|Ω)|[Rr](?=\s|$)|[kKmM](?:Ω|ohm)?))(?![A-Za-z])",
                    desc,
                    re.IGNORECASE,
                )
                if value_match:
                    passive_value = re.sub(r"\s+", "", value_match.group(1)).replace("µ", "u")
            spec = desc
            if name and name.lower() not in spec.lower():
                spec = f"{name} {spec}".strip()
            items.append({
                "name": passive_value or name or desc,
                "brand": value_at(row, brand_col),
                "package": value_at(row, package_col),
                "qty": _normalize_quantity(value_at(row, qty_col)),
                "spec": spec,
                "cat_key": cat_key,
                "subcat": subcat,
                "raw": raw,
            })
        return items

    @staticmethod
    def _detect_header_row(rows: list) -> int:
        """在前置标题/汇总行之后，自动定位最像字段名的表头行。"""
        header_tokens = {
            "订单", "料号", "商品", "型号", "名称", "品牌", "封装", "规格", "数量",
            "数量", "单位", "价格", "金额", "日期", "编号", "位号", "comment", "value",
            "part", "model", "value", "值", "reference", "ref", "footprint", "package",
            "manufacturer", "quantity", "qty", "description", "remarks", "comment",
        }
        best_index, best_score = 0, -1
        # 仅从前 20 行找表头，避免把后续真实数据误判为表头。
        for index, row in enumerate(rows[:20]):
            cells = [str(cell).strip() for cell in row if cell is not None and str(cell).strip()]
            if not cells:
                continue
            score = 0
            for cell in cells:
                normalized = cell.lower().replace(" ", "")
                if any(token in normalized for token in header_tokens):
                    score += 1
            # 表头通常至少有两个字段关键词；同分时取更靠前的一行。
            if score > best_score:
                best_index, best_score = index, score
        return best_index if best_score >= 2 else 0

    @staticmethod
    def _read_excel_rows(file_bytes: bytes, filename: str) -> list:
        """读取 .xlsx/.xls 为行数据；后缀与实际文件格式不一致时给出可理解的错误。"""
        suffix = os.path.splitext(filename)[1].lower()
        is_zip = file_bytes.startswith(b"PK\\x03\\x04")

        if suffix == ".xls" and not is_zip:
            try:
                book = xlrd.open_workbook(file_contents=file_bytes)
                candidates = []
                for index in range(book.nsheets):
                    sheet = book.sheet_by_index(index)
                    rows = [sheet.row_values(i) for i in range(sheet.nrows)]
                    candidates.append((BatchParser._sheet_header_score(rows), index, rows))
                return max(candidates, key=lambda item: (item[0], -item[1]))[2] if candidates else []
            except xlrd.biffh.XLRDError as e:
                raise ValueError("该文件不是可读取的 Excel 工作簿，请用 Excel/WPS 另存为 .xlsx 后再导入。") from e

        try:
            wb = load_workbook(BytesIO(file_bytes), data_only=True, read_only=True)
        except zipfile.BadZipFile as e:
            if suffix == ".xls":
                raise ValueError("文件扩展名是 .xls，但实际内容不是标准 Excel 文件。请用 Excel/WPS 打开后“另存为” .xlsx，再导入。") from e
            raise ValueError("该 .xlsx 文件内容不完整或并非真正的 Excel 文件。请用 Excel/WPS 打开并“另存为”新的 .xlsx 后重试。") from e
        try:
            candidates = []
            for index, ws in enumerate(wb.worksheets):
                rows = list(ws.iter_rows(values_only=True))
                candidates.append((BatchParser._sheet_header_score(rows), index, rows))
            return max(candidates, key=lambda item: (item[0], -item[1]))[2] if candidates else []
        finally:
            wb.close()

    @staticmethod
    def _sheet_header_score(rows: list) -> int:
        """为工作表评分，选择最像 BOM/采购明细的工作表。"""
        if not rows:
            return -1
        header_index = BatchParser._detect_header_row(rows)
        cells = [str(cell).strip().lower().replace(" ", "")
                 for cell in rows[header_index] if cell is not None and str(cell).strip()]
        tokens = (
            "型号", "商品型号", "model", "mpn", "value", "值", "数量", "qty",
            "quantity", "封装", "package", "footprint", "品牌", "brand", "manufacturer",
            "reference", "位号", "description", "描述",
        )
        return sum(1 for cell in cells if any(token in cell for token in tokens))

    def parse_excel(self, file_bytes: bytes, filename: str = "") -> tuple:
        """解析 Excel 文件。返回 (items, sheet_preview)。
        items 与 parse_text 同结构。sheet_preview 为前3行预览(用于前端展示)。"""
        rows = self._read_excel_rows(file_bytes, filename)
        if not rows:
            return [], []

        # 自动定位表头：兼容对账单/报表前的标题、合计、日期等非表格内容。
        header_index = self._detect_header_row(rows)
        header_row = rows[header_index]
        header = [str(c).strip() if c is not None else "" for c in header_row]
        data_rows = rows[header_index + 1:]
        preview = [header] + [list(r)[: min(len(header), 8)] for r in data_rows[:2]]

        # 采购明细、标准 BOM 等表格已具备型号/数量/描述字段时，优先纯规则直读。
        # 这样无需传输整表到远程模型；表头含义不明确时才回退 AI。
        structured_items = self._parse_structured_rows(header, data_rows)
        if structured_items is not None:
            self.dropped_nc = len(structured_items)
            structured_items = _fix_part_ref_name(structured_items)
            structured_items = _normalize_capacitor(structured_items)
            structured_items = _drop_nc_items(structured_items)
            self.dropped_nc -= len(structured_items)
            return structured_items, preview

        tree = _category_tree_text()
        items = []
        # 分批发送全部数据行 (每批 40 行) —— 旧版只发前5行样本导致最多识别5条
        for start in range(0, len(data_rows), 40):
            batch = data_rows[start:start + 40]
            prompt = f"""你是一个电子元器件 BOM 表分析助手。下面是一个 Excel 表格的内容:
表头: {json.dumps(header, ensure_ascii=False)}
本次数据行 (共 {len(batch)} 行, 全部列出):
{json.dumps([[str(c) if c is not None else "" for c in r] for r in batch], ensure_ascii=False)}

请分析每一列的含义, 并逐行解析每条元器件。输出 JSON:
{{"columns": {{"name": 列号, "brand": 列号, "package": 列号, "qty": 列号, "spec": 列号或null}},
 "items": [{{"row": 行号, "name": "...", "brand": "...", "package": "...", "qty": "...",
            "spec": "...", "cat": "一级分类名", "subcat": "子分类名"}}]}}
列号从 0 开始 (0 为第一列)。表头可能是 位号/型号/描述/数量/封装/规格/品牌 等任意叫法, 按实际含义识别。
分类体系 (一级分类: 子分类):
{tree}
找不到对应列的字段填空, qty 默认 10。
name 应填型号或规格值(如 10uF、10kΩ、STM32F103C8T6); 若某列只是位号编号(如 C1/R2/L3/U5 这种 字母+数字), 不要把它当 name, 用规格/值列作为 name。
电容: name 只填容值(如 10uF、100nF), 耐压(如 50V、25V)和材质(如 X7R、X5R、C0G、电解、钽)必须提取进 spec, 格式如 "50V X7R"。
必须解析本批全部 {len(batch)} 行, 一条都不能漏。
只输出紧凑 JSON(不要缩进换行), 不要多余文字。"""

            raw = self._chat([{"role": "user", "content": prompt}])
            data = self._parse_json(raw)
            for d in data.get("items", []):
                if not isinstance(d, dict):
                    continue
                cat_key, subcat = _match_cat_subcat(
                    str(d.get("cat", "")), str(d.get("subcat", ""))
                )
                # raw: 原始行拼接 (位号编号兜底用; row 可能 0/1 基, 两种都试)
                raw_txt = ""
                try:
                    ridx = int(d.get("row", -1))
                    for cand in (ridx, ridx - 1):
                        if 0 <= cand < len(batch):
                            raw_txt = " ".join(
                                str(c) for c in batch[cand] if c is not None
                            )
                            break
                except (TypeError, ValueError):
                    pass
                items.append({
                    "name": str(d.get("name", "")).strip(),
                    "brand": str(d.get("brand", "")).strip(),
                    "package": str(d.get("package", "")).strip(),
                    "qty": _normalize_quantity(d.get("qty", "10")),
                    "spec": str(d.get("spec", "")).strip(),
                    "cat_key": cat_key,
                    "subcat": subcat,
                    "raw": raw_txt,
                })
        self.dropped_nc = len(items)
        items = _fix_part_ref_name(items)
        items = _normalize_capacitor(items)
        items = _drop_nc_items(items)
        self.dropped_nc -= len(items)
        return items, preview

    # ── 批量写入 ────────────────────────────────────────
    def commit(self, items: list, data_dir: str) -> dict:
        """按子分类分组写入 Excel。items: [{name,brand,package,qty,spec,cat_key,subcat}]。
        返回 {subcat: 写入条数}。"""
        from warehouse.excel_store import ExcelStore
        from warehouse.activity import record as record_activity

        store = ExcelStore(data_dir)
        headers = [label for _, label in COMMON_FIELDS]

        # 按 (cat_key, subcat) 分组
        groups = {}
        for it in items:
            key = (it.get("cat_key"), it.get("subcat", ""))
            groups.setdefault(key, []).append(it)

        result = {}
        for (cat_key, subcat), grp in groups.items():
            if cat_key is None:
                # 未识别分类: 写入「未分类」区, 用户可手动归类, 不再丢弃
                from warehouse import unclassified
                from warehouse import undo as undo_mod
                h_u, old_uncat = unclassified.load(data_dir)
                unclassified.add(data_dir, grp)
                undo_mod.push(data_dir, [
                    {"subcat": undo_mod.UNCAT_KEY, "old_rows": old_uncat},
                ], "批量导入(未分类)")
                result["未分类"] = result.get("未分类", 0) + len(grp)
                continue
            # 读取现有数据, 与新增合并 (相同元件数量累加, 不产生重复行)
            old_headers, old_rows = store.load(subcat) if subcat else ([], [])
            new_rows = _merge_rows(old_rows, grp, subcat)
            path = store.save(subcat, headers, new_rows)
            record_activity(data_dir, subcat, old_rows, new_rows, path=path)
            from warehouse import undo as undo_mod
            undo_mod.push(data_dir, [{"subcat": subcat, "old_rows": old_rows}], "批量导入")
            result[subcat or "(未分类)"] = len(grp)
        return result


def parse_ai_config(app_dir: str) -> dict:
    """读取 AI 配置, 未配置返回 None。"""
    s = load_settings(app_dir)
    if not s["ai"].get("api_key"):
        return None
    return s["ai"]
