# -*- coding: utf-8 -*-
"""元器件仓库 — 数据包导出/导入 (ZIP)。

导出: 所有分类 xlsx + 未分类 + manifest.json → data/exports/parts-warehouse_导出_<时间戳>.zip
      排除运行时数据 (cache/backgrounds/decor/日志/设置/旧导出包)。
导入: 合并模式, 同名子分类按四键规则合并数量, 导入前自动备份到 backups/。
"""
import datetime
import io
import json
import os
import re
import shutil
import zipfile

from warehouse.config import COMMON_FIELDS

# 导出一级分类目录时排除的运行时目录/文件
RUNTIME_NAMES = {"cache", "backgrounds", "decor", "exports", "backups",
                 "settings.json", "activity_log.jsonl", "undo_log.jsonl",
                 "未分类"}

EXPORT_SUBDIR = "exports"
BACKUP_ROOT_REL = "backups"     # 项目根下的 backups/ (与 pack.py 一致)


def _category_dirs(data_dir: str) -> list:
    """data 下一级分类目录 (排除运行时目录)。"""
    dirs = []
    for name in sorted(os.listdir(data_dir)):
        p = os.path.join(data_dir, name)
        if os.path.isdir(p) and name not in RUNTIME_NAMES:
            dirs.append(p)
    return dirs


def _count_items(xlsx_path: str) -> int:
    """统计 xlsx 数据行数。"""
    try:
        from openpyxl import load_workbook
        wb = load_workbook(xlsx_path, read_only=True, data_only=True)
        ws = wb.active
        n = 0
        for row in ws.iter_rows(min_row=2, values_only=True):
            if any(c is not None and str(c).strip() for c in row):
                n += 1
        wb.close()
        return n
    except Exception:
        return 0


def export_package(data_dir: str, out_dir: str = "") -> dict:
    """打包全部元器件数据为 ZIP。

    out_dir 为空 → data/exports/; 否则导出到指定目录。
    返回 {ok, path, filename, subcats, items}。
    """
    from warehouse import unclassified
    export_dir = out_dir.strip() or os.path.join(data_dir, EXPORT_SUBDIR)
    if not os.path.isdir(export_dir):
        raise ValueError(f"导出目录不存在: {export_dir}")
    os.makedirs(export_dir, exist_ok=True)
    ts = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"parts-warehouse_导出_{ts}.zip"
    out_path = os.path.join(export_dir, filename)

    subcats = 0
    items = 0
    buf = io.BytesIO()
    with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as zf:
        for cat_dir in _category_dirs(data_dir):
            cat_name = os.path.basename(cat_dir)
            for fname in sorted(os.listdir(cat_dir)):
                if not fname.lower().endswith(".xlsx"):
                    continue
                fp = os.path.join(cat_dir, fname)
                n = _count_items(fp)
                if n == 0:
                    continue          # 空分类不打包
                zf.write(fp, os.path.join(cat_name, fname))
                subcats += 1
                items += n
        # 未分类
        h_u, rows = unclassified.load(data_dir)
        if rows:
            uncat_dir = os.path.join(data_dir, "未分类")
            os.makedirs(uncat_dir, exist_ok=True)
            uncat_path = os.path.join(uncat_dir, "未分类.xlsx")
            from openpyxl import Workbook
            wb = Workbook()
            ws = wb.active
            ws.append([label for _, label in COMMON_FIELDS])
            for r in rows:
                ws.append(r)
            wb.save(uncat_path)
            zf.write(uncat_path, "未分类/未分类.xlsx")
            subcats += 1
            items += len(rows)
        # manifest
        manifest = {
            "app": "元器件仓库", "format_version": 1,
            "exported_at": ts, "subcats": subcats, "items": items,
        }
        zf.writestr("manifest.json", json.dumps(manifest, ensure_ascii=False, indent=2))

    with open(out_path, "wb") as f:
        f.write(buf.getvalue())
    return {"ok": True, "path": out_path, "filename": filename,
            "subcats": subcats, "items": items}


def _extract_safe(zf: zipfile.ZipFile, tmp_dir: str):
    """安全解压 (防路径穿越)。返回 xlsx 文件相对路径列表。"""
    files = []
    for info in zf.infolist():
        name = info.filename.replace("\\", "/")
        if name.startswith("/") or ".." in name.split("/"):
            raise ValueError(f"数据包包含非法路径: {name}")
        target = os.path.join(tmp_dir, name)
        if not os.path.abspath(target).startswith(os.path.abspath(tmp_dir)):
            raise ValueError(f"数据包路径越界: {name}")
        if info.is_dir():
            os.makedirs(target, exist_ok=True)
            continue
        os.makedirs(os.path.dirname(target), exist_ok=True)
        with zf.open(info) as src, open(target, "wb") as dst:
            shutil.copyfileobj(src, dst)
        if name.lower().endswith(".xlsx"):
            files.append(name)
    return files


def _norm_subcat(name: str) -> str:
    """子分类名归一化: 去空格/连字符, 斜杠去除, 小写。用于旧名→标准名匹配。"""
    return re.sub(r"[\s\-—/]+", "", name or "").lower()


def _find_subcat(name: str) -> str | None:
    """在分类体系里找子分类: 精确 > 归一化模糊。"""
    from warehouse.config import CATEGORIES
    for cats in CATEGORIES.values():
        for sc in cats[1]:
            if sc == name:
                return sc
    n = _norm_subcat(name)
    for cats in CATEGORIES.values():
        for sc in cats[1]:
            if _norm_subcat(sc) == n:
                return sc
    return None


def import_package(data_dir: str, zip_bytes: bytes, project_root: str = "") -> dict:
    """合并导入 ZIP 数据包。

    返回 {ok, subcats, items, backup, detail: {子分类: 导入条数}}。
    """
    import tempfile
    from openpyxl import load_workbook
    from warehouse.excel_store import ExcelStore
    from warehouse.batch_import import _merge_rows
    from warehouse.activity import record as record_activity
    from warehouse import unclassified
    from warehouse import undo as undo_mod

    # 1. 解压校验
    with zipfile.ZipFile(io.BytesIO(zip_bytes)) as zf:
        names = zf.namelist()
        if "manifest.json" not in names:
            # 兼容无 manifest 的包: 只要有 xlsx 就接受
            has_xlsx = any(n.lower().endswith(".xlsx") for n in names)
            if not has_xlsx:
                raise ValueError("不是有效的数据包: 未找到 manifest.json 或 xlsx 文件")
        with tempfile.TemporaryDirectory() as tmp:
            files = _extract_safe(zf, tmp)

            # 2. 导入前备份
            backup = ""
            if project_root:
                bk_root = os.path.join(project_root, BACKUP_ROOT_REL)
                os.makedirs(bk_root, exist_ok=True)
                ts = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
                backup = os.path.join(bk_root, f"parts-warehouse_data_{ts}")
                shutil.copytree(data_dir, backup)

            store = ExcelStore(data_dir)
            detail = {}
            total_items = 0
            for rel in files:
                fp = os.path.join(tmp, rel)
                parts = rel.split("/")
                fname = os.path.basename(rel)
                # 未分类
                if fname == "未分类.xlsx" or (len(parts) >= 2 and parts[0] == "未分类"):
                    h_u, old_uncat = unclassified.load(data_dir)
                    from openpyxl import load_workbook
                    wb = load_workbook(fp, data_only=True)
                    ws = wb.active
                    grp = []
                    for row in ws.iter_rows(min_row=2, values_only=True):
                        if any(c is not None and str(c).strip() for c in row):
                            grp.append(list(row))
                    wb.close()
                    if grp:
                        unclassified.add(data_dir, grp)
                        undo_mod.push(data_dir, [
                            {"subcat": undo_mod.UNCAT_KEY, "old_rows": old_uncat},
                        ], "导入数据包(未分类)")
                    detail["未分类"] = detail.get("未分类", 0) + len(grp)
                    total_items += len(grp)
                    continue
                # 常规子分类: 文件名 = 子分类名
                subcat = os.path.splitext(fname)[0]
                try:
                    old_headers, old_rows = store.load(subcat)
                except ValueError:
                    # 旧命名 (如 "TVS-ESD" vs 标准 "TVS / ESD"): 归一化匹配标准名
                    std = _find_subcat(subcat)
                    if std:
                        subcat = std
                        try:
                            old_headers, old_rows = store.load(subcat)
                        except ValueError:
                            old_headers, old_rows = [], []
                    else:
                        # 真未知分类: 数据进未分类区, 备注保留原分类名
                        h_u, old_uncat = unclassified.load(data_dir)
                        wb = load_workbook(fp, data_only=True)
                        ws = wb.active
                        grp = []
                        for row in ws.iter_rows(min_row=2, values_only=True):
                            if any(c is not None and str(c).strip() for c in row):
                                r = list(row)
                                r = (r + [""] * 10)[:10]
                                grp.append({
                                    "name": str(r[0] or ""), "brand": str(r[1] or ""),
                                    "package": str(r[2] or ""), "qty": str(r[3] or "10"),
                                    "spec": str(r[6] if len(r) > 6 else ""),
                                    "raw": f"[原分类:{subcat}]",
                                })
                        wb.close()
                        if grp:
                            unclassified.add(data_dir, grp)
                            undo_mod.push(data_dir, [
                                {"subcat": undo_mod.UNCAT_KEY, "old_rows": old_uncat},
                            ], "导入数据包(未分类)")
                        detail[f"{subcat}→未分类"] = len(grp)
                        total_items += len(grp)
                        continue
                try:
                    headers = [label for _, label in COMMON_FIELDS]
                    wb = load_workbook(fp, data_only=True)
                    ws = wb.active
                    new_rows = []
                    for row in ws.iter_rows(min_row=2, values_only=True):
                        if any(c is not None and str(c).strip() for c in row):
                            new_rows.append(list(row))
                    wb.close()
                    if not new_rows:
                        continue
                    # 包内行(list) 转 dict, 供 _merge_rows 合并 (四键去重+数量累加)
                    new_items = [{
                        "name": str(r[0] or ""), "brand": str(r[1] or ""),
                        "package": str(r[2] or ""), "qty": str(r[3] or "10"),
                        "spec": str(r[6] if len(r) > 6 else ""),
                    } for r in new_rows]
                    merged = _merge_rows(old_rows, new_items, subcat)
                    path = store.save(subcat, headers, merged)
                    record_activity(data_dir, subcat, old_rows, merged, path=path)
                    undo_mod.push(data_dir, [{"subcat": subcat, "old_rows": old_rows}],
                                  "导入数据包")
                    added = len(merged) - len(old_rows)
                    detail[subcat] = added
                    total_items += added
                except Exception as e:
                    detail[f"{subcat}(跳过)"] = str(e)[:60]
            return {"ok": True, "subcats": len(detail), "items": total_items,
                    "backup": backup, "detail": detail}
