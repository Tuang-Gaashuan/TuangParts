# -*- coding: utf-8 -*-
"""元器件仓库 — Web 服务 (三级结构: 一级分类 -> 子分类 -> 元器件表格).

结构:
  总览页    : 一级分类卡片 (只显示有记录的分类)
  子分类页  : 一级分类下的子分类列表 (只显示有数据的子分类)
  表格页    : 子分类的元器件 Excel (可编辑 / AI 填入 / 保存)

数据文件: data/<一级分类>/<子分类>.xlsx, 有数据才建文件。
重名子分类物理文件唯一, 多个一级分类入口指向同一文件。

设置: data/settings.json (数据路径 / AI 接口 / 界面主题), 用户可改。

启动:  python app.py   (浏览器 http://127.0.0.1:5000)
桌面版: python desktop.py (pywebview 原生窗口)
"""

import os
import threading
import webbrowser
import faulthandler
from datetime import datetime

# 原生层崩溃 (OpenCV/摄像头驱动) 无 Python traceback, faulthandler 能打印线程栈定位
faulthandler.enable()

from flask import Flask, jsonify, render_template, request, send_from_directory, send_file

from warehouse.config import CATEGORIES, fields_for, subcats, subcat_owners, primary_owner, safe_filename
from warehouse.excel_store import ExcelStore
from warehouse.ai_fill import AIFiller, get_api_key
from warehouse.settings import load_settings, save_settings, public_view, resolve_data_dir, chat_completions_url
from warehouse.activity import record as record_activity
from warehouse.activity import load as load_activity
from warehouse.batch_import import BatchParser

# 数据/设置根目录:
#   打包版 (desktop.py) 通过环境变量 PARTS_APP_DIR 指向 exe 旁目录;
#   开发版用源码目录。
BASE_DIR = os.environ.get("PARTS_APP_DIR") or os.path.dirname(os.path.abspath(__file__))
DEFAULT_DATA_DIR = os.path.join(BASE_DIR, "data")

# 模板/静态资源目录: 打包版在 _MEIPASS (通过 PARTS_RES_DIR), 开发版用源码目录
RES_DIR = os.environ.get("PARTS_RES_DIR") or BASE_DIR

app = Flask(
    __name__,
    template_folder=os.path.join(RES_DIR, "templates"),
    static_folder=os.path.join(RES_DIR, "static"),
)


def get_data_dir() -> str:
    """当前实际数据目录 (跟随用户设置, 支持运行时切换)。"""
    return resolve_data_dir(BASE_DIR)


def get_store() -> ExcelStore:
    return app.config.get("STORE") or ExcelStore(get_data_dir())


def get_ai_cfg() -> dict | None:
    """当前 AI 配置, 未配置返回 None。

    online: 需要 api_key; ollama: 本地离线模型, 无需 key,
    默认 base_url http://localhost:11434/v1 (OpenAI 兼容), key 用占位符。
    """
    s = load_settings(BASE_DIR)
    ai = s["ai"]
    provider = ai.get("provider", "online")
    if provider == "ollama":
        cfg = dict(ai)
        cfg["base_url"] = (ai.get("base_url") or "http://localhost:11434/v1").strip()
        cfg["api_key"] = "ollama"   # 本地服务不校验, 占位即可
        if not cfg.get("model"):
            cfg["model"] = "qwen2.5:7b"
        return cfg
    if not ai.get("api_key"):
        return None
    return ai


# ── 页面 ───────────────────────────────────────────────
@app.route("/")
def index():
    return render_template("index.html")


# ── API: 设置 (数据路径 / AI / 主题) ────────────────────
@app.route("/api/settings", methods=["GET"])
def settings_get():
    s = load_settings(BASE_DIR)
    view = public_view(s)
    view["resolved_data_dir"] = get_data_dir()
    return jsonify(view)


@app.route("/api/settings", methods=["POST"])
def settings_post():
    data = request.get_json(force=True) or {}
    patch = {}
    if "data_dir" in data:
        patch["data_dir"] = (data.get("data_dir") or "").strip()
    if "theme" in data:
        patch["theme"] = data.get("theme", "light-blue")
    if "background" in data:
        patch["background"] = (data.get("background") or "").strip()
    if "bg_fill" in data:
        v = (data.get("bg_fill") or "cover").strip()
        if v in ("cover", "contain", "stretch", "repeat", "original"):
            patch["bg_fill"] = v
    for k in ("bg_brightness", "bg_opacity", "card_opacity", "decor1", "decor2"):
        if k in data:
            if k.startswith("decor"):
                patch[k] = (data.get(k) or "").strip()
            else:
                try:
                    patch[k] = max(0, min(200, int(data[k])))
                except (TypeError, ValueError):
                    pass
    if "ai" in data:
        new_ai = {
            "provider": (data["ai"].get("provider") or "online").strip(),
            "base_url": (data["ai"].get("base_url") or "").strip(),
            "api_key": (data["ai"].get("api_key") or "").strip(),
            "model": (data["ai"].get("model") or "").strip(),
        }
        # 留空 Key = 保持原有 Key (避免打开设置面板再点保存时误清空)
        cur_ai = load_settings(BASE_DIR)["ai"]
        if not new_ai["api_key"] and cur_ai.get("api_key") and new_ai["provider"] != "ollama":
            new_ai["api_key"] = cur_ai["api_key"]
        patch["ai"] = new_ai
    s = save_settings(BASE_DIR, patch)
    view = public_view(s)
    view["resolved_data_dir"] = get_data_dir()
    return jsonify(view)


@app.route("/api/settings/background", methods=["POST"])
def settings_background():
    """上传背景图, 只存文件不写设置 (预览用), 保存由前端统一点"保存背景设置"。"""
    if "file" not in request.files:
        return jsonify({"error": "未收到文件"}), 400
    f = request.files["file"]
    fname = f.filename or ""
    if not fname.lower().endswith((".png", ".jpg", ".jpeg", ".webp", ".bmp")):
        return jsonify({"error": "仅支持 png/jpg/webp/bmp 图片"}), 400
    try:
        data_dir = get_data_dir()
        bg_dir = os.path.join(data_dir, "backgrounds")
        os.makedirs(bg_dir, exist_ok=True)
        ext = os.path.splitext(fname)[1].lower()
        out_name = "bg_" + datetime.now().strftime("%Y%m%d_%H%M%S") + ext
        out_path = os.path.join(bg_dir, out_name)
        f.save(out_path)
        rel = os.path.join("backgrounds", out_name).replace("\\", "/")
        return jsonify({"ok": True, "background": rel})
    except Exception as e:
        return jsonify({"error": str(e)[:200]}), 500


@app.route("/api/settings/decor", methods=["POST"])
def settings_decor():
    """上传首页装饰图 (index=1|2), 存 data/decor/, 立即写设置。"""
    f = request.files.get("file")
    index = request.form.get("index", "1")
    if not f or not (f.filename or "").lower().endswith((".png", ".jpg", ".jpeg", ".webp", ".bmp")):
        return jsonify({"ok": False, "error": "请选择图片文件 (png/jpg/webp/bmp)"})
    try:
        data_dir = get_data_dir()
        decor_dir = os.path.join(data_dir, "decor")
        os.makedirs(decor_dir, exist_ok=True)
        ext = os.path.splitext(f.filename)[1].lower()
        out_name = f"decor_{datetime.now().strftime('%Y%m%d_%H%M%S')}_{index}{ext}"
        f.save(os.path.join(decor_dir, out_name))
        rel = f"decor/{out_name}"
        key = f"decor{index}"
        save_settings(BASE_DIR, {key: rel})
        return jsonify({"ok": True, "decor": rel})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)[:200]})


# ── API: data 目录静态服务 (背景图等) ──────────────────
@app.route("/data/<path:filename>")
def data_file(filename):
    return send_from_directory(get_data_dir(), filename)


# ── API: 总览 (一级分类, 只含有记录的) ─────────────────
@app.route("/api/overview")
def overview():
    store = get_store()
    overview = store.all_overview()
    cards = []
    for key, total in overview.items():
        if total > 0:
            name = "未分类" if key == "unclassified" else CATEGORIES[key][0]
            cards.append({"key": key, "name": name, "count": total})
    return jsonify({"cards": cards, "total": sum(overview.values())})


# ── API: 未分类元件 (手动归类界面) ─────────────────────
@app.route("/api/unclassified")
def unclassified_list():
    from warehouse import unclassified as uncat
    data_dir = get_data_dir()
    headers, rows = uncat.load(data_dir)
    fkeys = [fk for fk, _ in fields_for("")]
    items = []
    for i, row in enumerate(rows):
        d = {}
        for j, fk in enumerate(fkeys):
            v = row[j] if j < len(row) else ""
            d[fk] = v if v is not None else ""
        d["raw"] = row[len(fkeys)] if len(row) > len(fkeys) else ""
        d["index"] = i
        items.append(d)
    tree = {key: {"name": CATEGORIES[key][0], "subs": CATEGORIES[key][1]} for key in CATEGORIES}
    return jsonify({"items": items, "count": len(items), "cat_tree": tree})


@app.route("/api/unclassified/assign", methods=["POST"])
def unclassified_assign():
    data = request.get_json(force=True) or {}
    indices = data.get("indices") or []
    cat_key = (data.get("cat_key") or "").strip()
    subcat = (data.get("subcat") or "").strip()
    if not indices or cat_key not in CATEGORIES:
        return jsonify({"error": "请选择一级分类与条目"}), 400
    subs = CATEGORIES[cat_key][1]
    if subcat not in subs:
        return jsonify({"error": f"子分类不存在: {subcat}"}), 400
    from warehouse import unclassified as uncat
    data_dir = get_data_dir()
    try:
        # 记录撤销快照 (目标子分类 + 未分类)
        headers_u, old_uncat_rows = uncat.load(data_dir)
        h_old, old_target_rows = get_store().load(subcat)
        moved = uncat.assign(data_dir, indices, cat_key, subcat)
    except Exception as e:
        return jsonify({"error": str(e)[:200]}), 500
    from warehouse import undo as undo_mod
    undo_mod.push(data_dir, [
        {"subcat": subcat, "old_rows": old_target_rows},
        {"subcat": undo_mod.UNCAT_KEY, "old_rows": old_uncat_rows},
    ], "未分类归类")
    return jsonify({"ok": True, "moved": moved})


# ── API: 一级分类下的子分类 (只含有数据的) ──────────────
@app.route("/api/category/<key>")
def category(key):
    if key not in CATEGORIES:
        return jsonify({"error": "未知分类"}), 404
    store = get_store()
    subs = store.subcats_with_data(key)
    return jsonify({
        "key": key,
        "name": CATEGORIES[key][0],
        "subcats": [{"name": name, "count": n, "owner": primary_owner(name)} for name, n in subs],
    })


# ── API: 子分类表格数据 ────────────────────────────────
@app.route("/api/subcat")
def subcat():
    name = (request.args.get("name") or "").strip()
    if not name or primary_owner(name) is None:
        return jsonify({"error": "未知子分类"}), 404
    store = get_store()
    headers, rows = store.load(name)
    fields = fields_for(name)
    items = []
    for row in rows:
        item = {}
        for (fkey, label), val in zip(fields, row):
            item[fkey] = val if val is not None else ""
        items.append(item)
    return jsonify({
        "name": name,
        "owners": subcat_owners(name),
        "fields": [{"key": fk, "label": lb} for fk, lb in fields],
        "items": items,
    })


# ── API: 保存子分类 ────────────────────────────────────
@app.route("/api/save", methods=["POST"])
def save():
    data = request.get_json(force=True)
    name = (data.get("name") or "").strip()
    if not name or primary_owner(name) is None:
        return jsonify({"error": "未知子分类"}), 404
    store = get_store()
    data_dir = get_data_dir()
    fields = fields_for(name)
    fkeys = [fk for fk, _ in fields]
    items = data.get("items", [])
    rows = []
    for item in items:
        row = [item.get(fk, "") for fk in fkeys]
        rows.append(row)
    old_headers, old_rows = store.load(name)
    path = store.save(name, [lb for _, lb in fields], rows)
    record_activity(data_dir, name, old_rows, rows, path=os.path.relpath(path, BASE_DIR))
    from warehouse import undo as undo_mod
    undo_mod.push(data_dir, [{"subcat": name, "old_rows": old_rows}], "保存修改")
    return jsonify({"ok": True, "saved": len(rows), "path": os.path.relpath(path, BASE_DIR)})


# ── API: 一级分类下的全量子分类 (含空的, 录入界面用) ──
@app.route("/api/subcats_all")
def subcats_all():
    key = request.args.get("key", "")
    if key not in CATEGORIES:
        return jsonify({"error": "未知分类"}), 404
    return jsonify({"key": key, "subcats": subcats(key)})


# ── API: 取出元器件 (出库) ─────────────────────────────
@app.route("/api/withdraw", methods=["POST"])
def withdraw():
    """从子分类扣减库存。items: [{row: 行号, qty: 取出数量}]。
    校验数量充足后扣减, 记录操作日志 (qty_delta<0 → 使用/取出)。"""
    data = request.get_json(force=True) or {}
    name = (data.get("name") or "").strip()
    items = data.get("items") or []
    if not name or primary_owner(name) is None:
        return jsonify({"error": "未知子分类"}), 404
    if not items:
        return jsonify({"error": "未选择要取出的元件"}), 400

    store = get_store()
    data_dir = get_data_dir()
    headers, old_rows = store.load(name)

    # 校验并计算新行
    new_rows = [list(r) for r in old_rows]
    taken_total = 0
    detail = []
    for it in items:
        try:
            row_index = int(it.get("row", -1))
            take = int(float(str(it.get("qty", 0)).strip()))
        except (TypeError, ValueError):
            continue
        if row_index < 0 or row_index >= len(new_rows):
            continue
        if take <= 0:
            continue
        row = new_rows[row_index]
        try:
            cur = int(float(str(row[3]).strip())) if len(row) > 3 else 0
        except (TypeError, ValueError):
            cur = 0
        if take > cur:
            label = str(row[0]) if row and row[0] else f"第{row_index + 1}行"
            return jsonify({"error": f"「{label}」库存不足: 当前 {cur} 件, 要取出 {take} 件"}), 400
        row[3] = str(cur - take)
        taken_total += take
        detail.append({"row": row_index, "name": row[0] if row else "", "taken": take})

    if taken_total == 0:
        return jsonify({"error": "取出数量必须大于 0"}), 400

    path = store.save(name, [label for _, label in fields_for(name)], new_rows)
    record_activity(data_dir, name, old_rows, new_rows, path=os.path.relpath(path, BASE_DIR))
    from warehouse import undo as undo_mod
    undo_mod.push(data_dir, [{"subcat": name, "old_rows": old_rows}], "取出")
    return jsonify({"ok": True, "subcat": name, "taken": taken_total, "items": detail})


# ── API: BOM 取出匹配 (替换料机制) ─────────────────────
@app.route("/api/withdraw/match", methods=["POST"])
def withdraw_match():
    """对 BOM 解析结果匹配现有库存。
    items: 批量导入解析后的条目 [{name,brand,package,qty,spec,cat_key,subcat}]。
    返回 [{item, candidates: [{subcat,row,name,brand,package,qty,spec}]}]。
    """
    data = request.get_json(force=True) or {}
    items = data.get("items") or []
    if not items:
        return jsonify({"error": "无数据"}), 400
    from warehouse.withdraw_match import match_items
    results = match_items(items, get_data_dir())
    matched = sum(1 for r in results if r["candidates"])
    return jsonify({"ok": True, "results": results, "matched": matched, "total": len(results)})


@app.route("/api/addstock", methods=["POST"])
def addstock():
    """给子分类元件增加库存。items: [{row: 行号, qty: 增加数量}]。
    数量累加到库存, 记录操作日志 (qty_delta>0 → 存入)。"""
    data = request.get_json(force=True) or {}
    name = (data.get("name") or "").strip()
    items = data.get("items") or []
    if not name or primary_owner(name) is None:
        return jsonify({"error": "未知子分类"}), 404
    if not items:
        return jsonify({"error": "未选择要增加的元件"}), 400

    store = get_store()
    data_dir = get_data_dir()
    headers, old_rows = store.load(name)
    new_rows = [list(r) for r in old_rows]
    added_total = 0
    detail = []
    for it in items:
        try:
            row_index = int(it.get("row", -1))
            add = int(float(str(it.get("qty", 0)).strip()))
        except (TypeError, ValueError):
            continue
        if row_index < 0 or row_index >= len(new_rows) or add <= 0:
            continue
        row = new_rows[row_index]
        try:
            cur = int(float(str(row[3]).strip())) if len(row) > 3 else 0
        except (TypeError, ValueError):
            cur = 0
        row[3] = str(cur + add)
        added_total += add
        detail.append({"row": row_index, "name": row[0] if row else "", "added": add})

    if added_total == 0:
        return jsonify({"error": "增加数量必须大于 0"}), 400

    path = store.save(name, [label for _, label in fields_for(name)], new_rows)
    record_activity(data_dir, name, old_rows, new_rows, path=os.path.relpath(path, BASE_DIR))
    from warehouse import undo as undo_mod
    undo_mod.push(data_dir, [{"subcat": name, "old_rows": old_rows}], "增加库存")
    return jsonify({"ok": True, "subcat": name, "added": added_total, "items": detail})


# ── API: 合并重复元件 (名称/品牌/封装/规格 相同 → 数量累加) ──
@app.route("/api/subcat/merge", methods=["POST"])
def subcat_merge():
    data = request.get_json(force=True) or {}
    name = (data.get("name") or "").strip()
    if not name or primary_owner(name) is None:
        return jsonify({"error": "未知子分类"}), 404
    store = get_store()
    data_dir = get_data_dir()
    headers, old_rows = store.load(name)
    from warehouse.batch_import import _merge_rows
    new_rows = _merge_rows(old_rows, [], name)
    removed = len(old_rows) - len(new_rows)
    if removed <= 0:
        return jsonify({"ok": True, "removed": 0, "before": len(old_rows), "after": len(new_rows)})
    path = store.save(name, [label for _, label in fields_for(name)], new_rows)
    record_activity(data_dir, name, old_rows, new_rows, path=os.path.relpath(path, BASE_DIR))
    from warehouse import undo as undo_mod
    undo_mod.push(data_dir, [{"subcat": name, "old_rows": old_rows}], "合并重复")
    return jsonify({"ok": True, "removed": removed, "before": len(old_rows), "after": len(new_rows)})


# ── API: 快速删除数据 (设置 → 数据管理) ────────────────
def _clear_all_data(data_dir: str) -> dict:
    """清空全部元器件数据: 分类文件夹 + 未分类 + 活动日志。
    保留: settings.json / backgrounds / decor (配置与图片资源)。"""
    import shutil
    cleared = {"categories": [], "activity": False}
    for name in os.listdir(data_dir):
        p = os.path.join(data_dir, name)
        if name in ("settings.json", "backgrounds", "decor"):
            continue
        if os.path.isdir(p):
            shutil.rmtree(p)
            cleared["categories"].append(name)
        elif name == "activity_log.jsonl":
            os.remove(p)
            cleared["activity"] = True
        else:
            os.remove(p)
    return cleared


@app.route("/api/data/clear", methods=["POST"])
def data_clear():
    data = request.get_json(force=True) or {}
    scope = data.get("scope", "")
    data_dir = get_data_dir()
    if scope == "all":
        cleared = _clear_all_data(data_dir)
        return jsonify({"ok": True, "scope": "all", "cleared": cleared})
    if scope == "unclassified":
        from warehouse import unclassified
        n = unclassified.count(data_dir)
        if n:
            unclassified.remove(data_dir, list(range(n)))
        return jsonify({"ok": True, "scope": "unclassified", "cleared": n})
    if scope == "activity":
        p = os.path.join(data_dir, "activity_log.jsonl")
        if os.path.exists(p):
            os.remove(p)
        return jsonify({"ok": True, "scope": "activity"})
    return jsonify({"error": "未知操作"}), 400


# ── API: 删除指定子分类 (设置 → 数据管理) ──────────────
@app.route("/api/subcat/delete", methods=["POST"])
def subcat_delete():
    data = request.get_json(force=True) or {}
    names = data.get("names") or []
    if not names:
        return jsonify({"error": "未选择子分类"}), 400
    store = get_store()
    data_dir = get_data_dir()
    deleted = []
    snapshots = []
    for name in names:
        if primary_owner(name) is None:
            continue
        headers, old_rows = store.load(name)
        if not old_rows:
            continue
        path = store.save(name, [label for _, label in fields_for(name)], [])
        record_activity(data_dir, name, old_rows, [], path=os.path.relpath(path, BASE_DIR))
        deleted.append(name)
        snapshots.append({"subcat": name, "old_rows": old_rows})
    if snapshots:
        from warehouse import undo as undo_mod
        undo_mod.push(data_dir, snapshots, "删除子分类")
    return jsonify({"ok": True, "deleted": deleted})


# ── API: 撤回系统 ─────────────────────────────────────
@app.route("/api/undo", methods=["GET"])
def undo_list():
    from warehouse import undo as undo_mod
    return jsonify({"ok": True, "items": undo_mod.list_recent(get_data_dir())})


@app.route("/api/undo", methods=["POST"])
def undo_apply():
    data = request.get_json(force=True) or {}
    time = (data.get("time") or "").strip()
    if not time:
        return jsonify({"error": "缺少操作时间"}), 400
    from warehouse import undo as undo_mod
    entry = undo_mod.undo(get_data_dir(), time)
    if not entry:
        return jsonify({"error": "该操作不存在或已撤回"}), 404
    return jsonify({"ok": True, "action": entry.get("action", ""),
                    "restored": entry.get("restored", [])})


# ── API: 浏览页 (统计 + 日志 + 低库存) ─────────────────
@app.route("/api/dashboard")
def dashboard():
    store = get_store()
    stats = store.global_stats()
    logs = load_activity(get_data_dir(), limit=50)
    return jsonify({"stats": stats, "logs": logs})


@app.route("/api/lowstock")
def lowstock():
    store = get_store()
    threshold = request.args.get("threshold", default=10, type=int)
    items = store.low_stock(threshold)
    return jsonify({"threshold": threshold, "items": items})


# ── API: 品牌库 (采购参考: 品牌 / 采购量 / 经营的元器件类别) ──
@app.route("/api/brands")
def brands():
    """品牌聚合: 品牌名称 + 条目数 + 总数量 + 覆盖的一级分类/子分类 + 型号示例。"""
    from warehouse.brands import aggregate
    data = aggregate(get_data_dir())
    return jsonify(data)


@app.route("/api/brands/detail")
def brands_detail():
    """某品牌(含所有别名写法)下的全部元器件行 (只读明细表)。"""
    brand = (request.args.get("brand") or "").strip()
    if not brand:
        return jsonify({"error": "缺少品牌名"}), 400
    from warehouse.brands import detail as brands_detail_rows
    res = brands_detail_rows(get_data_dir(), brand)
    return jsonify({"brand": brand, "items": res["items"], "count": len(res["items"]),
                    "aliases": res["aliases"]})


@app.route("/api/brands/export")
def brands_export():
    """品牌汇总导出为 xlsx (直接下载)。"""
    from io import BytesIO
    from warehouse.brands import export_xlsx
    try:
        data = export_xlsx(get_data_dir())
    except Exception as e:
        return jsonify({"error": f"导出失败: {str(e)[:200]}"}), 500
    fname = f"parts-warehouse_品牌汇总_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return send_file(
        BytesIO(data), as_attachment=True, download_name=fname,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@app.route("/api/brands/catalog", methods=["POST"])
def brands_catalog():
    """重建品牌库档案 data/品牌库.xlsx (品牌/业务/已购种类/总件数)。"""
    from warehouse.brands import build_catalog
    try:
        p = build_catalog(get_data_dir())
        # 绝对路径返回 (数据目录可能在别的盘, relpath 会跨盘崩溃)
        return jsonify({"ok": True, "path": p})
    except Exception as e:
        return jsonify({"error": f"生成失败: {str(e)[:200]}"}), 500


# ── API: AI 连通性测试 ─────────────────────────────────
@app.route("/api/ollama/models", methods=["GET"])
def ollama_models():
    """探测本地 Ollama: 是否运行 + 已装模型列表。

    返回 {running, base_url, models: [{name, size_gb}], error?}
    """
    import httpx
    url = "http://localhost:11434/api/tags"
    try:
        resp = httpx.get(url, timeout=5)
        if resp.status_code != 200:
            return jsonify({"running": False, "models": [],
                            "error": f"Ollama 服务响应异常 HTTP {resp.status_code}"})
        data = resp.json()
        models = [{
            "name": m.get("name", ""),
            "size_gb": round((m.get("size") or 0) / 1024 ** 3, 2),
        } for m in data.get("models", [])]
        return jsonify({"running": True, "models": models, "base_url": "http://localhost:11434/v1"})
    except Exception as e:
        return jsonify({"running": False, "models": [],
                        "error": "未检测到 Ollama 服务：请先安装并启动 Ollama（ollama.com），"
                                 "再用 `ollama pull qwen2.5:7b` 拉取模型"}), 200


@app.route("/api/ai_test", methods=["POST"])
def ai_test():
    cfg = get_ai_cfg()
    if not cfg:
        return jsonify({"ok": False, "error": "未配置 AI 接口，请先填写 API Key"}), 200
    try:
        import httpx
        url = chat_completions_url(cfg['base_url'])
        resp = httpx.post(
            url,
            headers={"Authorization": f"Bearer {cfg['api_key']}", "Content-Type": "application/json"},
            json={
                "model": cfg["model"],
                "messages": [{"role": "user", "content": "回复OK两个字"}],
                "max_tokens": 10,
                "temperature": 0,
            },
            timeout=30,
        )
        if resp.status_code != 200:
            return jsonify({"ok": False, "error": f"HTTP {resp.status_code}: {resp.text[:200]}"}), 200
        reply = resp.json()["choices"][0]["message"]["content"].strip()
        return jsonify({"ok": True, "reply": reply})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)[:300]}), 200


# ── API: 图片文字识别 (OCR) ────────────────────────────
@app.route("/api/ocr", methods=["POST"])
def ocr_recognize():
    """上传图片, 识别其中的文字 (料袋/照片/截图), 返回文本行。"""
    if "file" not in request.files:
        return jsonify({"error": "未收到图片"}), 400
    f = request.files["file"]
    fname = f.filename or ""
    if not fname.lower().endswith((".png", ".jpg", ".jpeg", ".webp", ".bmp")):
        return jsonify({"error": "仅支持 png/jpg/webp/bmp 图片"}), 400
    try:
        from warehouse.ocr import recognize
        lines = recognize(f.read())
        return jsonify({"ok": True, "text": "\n".join(lines), "lines": lines})
    except Exception as e:
        return jsonify({"error": f"OCR 识别失败: {str(e)[:200]}"}), 500


@app.route("/api/ocr/format", methods=["POST"])
def ocr_format():
    """把 OCR 识别文本按料袋模板整理 (数量 型号 品牌 封装 电气参数 器件名称)。"""
    data = request.get_json(force=True) or {}
    text = (data.get("text") or "").strip()
    if not text:
        return jsonify({"error": "没有可整理的文本"}), 400
    cfg = get_ai_cfg()
    if not cfg:
        return jsonify({"error": "未配置 AI 接口，无法自动整理"}), 400
    try:
        from warehouse.ocr import format_text
        out = format_text(text, cfg)
        return jsonify({"ok": True, "text": out})
    except Exception as e:
        return jsonify({"error": f"整理失败: {str(e)[:200]}"}), 500


# ── API: 摄像头拍照识别 ────────────────────────────────
def _camera_names() -> list:
    """用 Windows PnP 枚举摄像头设备友好名, 顺序与 DirectShow 索引一致。

    仅用于给用户看"哪个摄像头是哪个", 失败返回 [] (名称缺失不影响功能)。
    """
    import subprocess
    ps = (
        "[Console]::OutputEncoding=[Text.Encoding]::UTF8;"
        "Get-PnpDevice -PresentOnly | Where-Object { $_.Class -eq 'Camera' -or "
        "($_.Class -eq 'Image' -and $_.FriendlyName -match 'cam') } | "
        "Sort-Object InstanceId | ForEach-Object { $_.FriendlyName }"
    )
    try:
        r = subprocess.run(
            ["powershell", "-NoProfile", "-Command", ps],
            capture_output=True, timeout=15, text=True,
            encoding="utf-8", errors="replace",
        )
        return [ln.strip() for ln in r.stdout.splitlines() if ln.strip()]
    except Exception:
        return []


_WIN_TITLE = "PartsWarehouse-Camera (SPACE=CAP BS=UNDO ENTER=DONE)"


def _sanitize_ascii(s) -> str:
    """OpenCV putText 只能画 ASCII, 中文设备名过滤成 '?'。"""
    return "".join(ch if 32 <= ord(ch) < 127 else "?" for ch in (s or ""))


def _pick_resolution(cap) -> tuple:
    """探测摄像头实际能出的最高分辨率, 返回 (w, h)。"""
    import time
    import cv2
    for w, h in [(1920, 1080), (1280, 1024), (1280, 720), (960, 540)]:
        cap.set(cv2.CAP_PROP_FRAME_WIDTH, w)
        cap.set(cv2.CAP_PROP_FRAME_HEIGHT, h)
        time.sleep(0.08)
        ret, frame = cap.read()
        if ret and frame is not None:
            return frame.shape[1], frame.shape[0]
    return 1280, 720


LUM_DARK = 60        # 低于此值视为太暗 (实测用户环境约 57, OCR 甜区约 80~180)
LUM_BRIGHT = 200     # 高于此值视为太亮


def _lum_of(frame) -> int:
    """感知亮度 0~255 (OpenCV GRAY = 0.299R+0.587G+0.114B)。"""
    import cv2
    return int(cv2.cvtColor(frame, cv2.COLOR_BGR2GRAY).mean())


def _draw_hud(frame, n_photos: int, device: int, device_name: str, res_text: str,
              lum: int = 128, warn_text: str = "", flash: bool = False):
    """给预览帧画半透明 HUD: 顶部信息条 / 拍摄数徽标 / 实时亮度 / 三分线 / 底部按键条。

    putText 只能画 ASCII → 设备名先过 _sanitize_ascii, 中文提示留在网页上。
    """
    import cv2
    import numpy as np
    _FONT = cv2.FONT_HERSHEY_SIMPLEX
    h, w = frame.shape[:2]
    out = frame.copy()

    # 取景三分线 (半透明细线, 最底层)
    overlay = out.copy()
    thin = max(1, w // 900)
    for i in (1, 2):
        cv2.line(overlay, (w * i // 3, 0), (w * i // 3, h), (255, 255, 255), thin)
        cv2.line(overlay, (0, h * i // 3), (w, h * i // 3), (255, 255, 255), thin)
    out = cv2.addWeighted(overlay, 0.16, out, 0.84, 0)

    # 顶部信息条 (深色半透明, 两行: 标题+徽标 / 设备+分辨率+亮度)
    bar_h = max(54, h // 12)
    overlay = out.copy()
    cv2.rectangle(overlay, (0, 0), (w, bar_h), (10, 16, 28), -1)
    out = cv2.addWeighted(overlay, 0.66, out, 0.34, 0)
    cv2.putText(out, "PARTS WAREHOUSE", (14, 23), _FONT, 0.62,
                (255, 180, 60), 2, cv2.LINE_AA)
    name = _sanitize_ascii(device_name) or f"Cam {device}"
    cv2.putText(out, f"{name}  |  {res_text}", (14, bar_h - 10), _FONT, 0.5,
                (190, 205, 225), 1, cv2.LINE_AA)

    # 右上角拍摄数徽标 (第一行, 绿点 + 数字)
    label = f"{n_photos} SHOT"
    (tw, _), _ = cv2.getTextSize(label, _FONT, 0.6, 2)
    bx = w - tw - 26
    cv2.circle(out, (bx - 9, 22), 6, (50, 210, 100), -1)
    cv2.putText(out, label, (bx, 30), _FONT, 0.6, (255, 255, 255), 2, cv2.LINE_AA)

    # 亮度 (第二行右侧, 按三档着色: 正常绿 / 太暗橙 / 太亮黄)
    if lum < LUM_DARK:
        lum_label, lum_color = f"LUM {lum} DARK - LIGHT?", (0, 165, 255)
    elif lum > LUM_BRIGHT:
        lum_label, lum_color = f"LUM {lum} BRIGHT - SHADE?", (0, 255, 255)
    else:
        lum_label, lum_color = f"LUM {lum} NORMAL", (140, 255, 140)
    (lw, _), _ = cv2.getTextSize(lum_label, _FONT, 0.55, 2)
    cv2.putText(out, lum_label, (w - lw - 16, bar_h - 10), _FONT, 0.55,
                lum_color, 2, cv2.LINE_AA)

    # 拍照亮度异常警告浮层 (画面中上方, 大橙字, 约 1.3s 自动消失)
    if warn_text:
        (ww, wh), _ = cv2.getTextSize(warn_text, _FONT, 1.0, 3)
        wx, wy = (w - ww) // 2, max(bar_h + 46, int(h * 0.2))
        overlay = out.copy()
        cv2.rectangle(overlay, (wx - 16, wy - wh - 10), (wx + ww + 16, wy + 8),
                      (10, 16, 28), -1)
        out = cv2.addWeighted(overlay, 0.7, out, 0.3, 0)
        cv2.putText(out, warn_text, (wx, wy), _FONT, 1.0, (0, 165, 255), 3, cv2.LINE_AA)

    # 底部按键条 (深色半透明)
    bar_h2 = max(52, h // 11)
    overlay = out.copy()
    cv2.rectangle(overlay, (0, h - bar_h2), (w, h), (10, 16, 28), -1)
    out = cv2.addWeighted(overlay, 0.66, out, 0.34, 0)
    cv2.putText(out, "[SPACE] CAPTURE   [BS] UNDO   [ENTER] DONE   [ESC] EXIT",
                (14, h - bar_h2 // 2 + 6), _FONT, 0.6, (255, 255, 255), 1, cv2.LINE_AA)

    # 拍照闪光反馈
    if flash:
        out = cv2.addWeighted(out, 0.76, np.full_like(out, 255), 0.24, 0)
    return out


def _camera_capture_session(data_dir: str, device: int = 1, max_photos: int = 30,
                            timeout_s: int = 180, device_name: str = "") -> list:
    """打开指定摄像头弹窗: 空格=拍照, 回车=结束。

    自动取该摄像头最高分辨率, 拍照存原帧 (OCR 识别更清晰), 预览带 HUD。
    照片存 data/cache/ 后返回 bytes 列表 (调用方识别后删除)。
    返回 [] 表示摄像头不可用或未拍到。
    """
    import time
    import cv2
    import numpy as np

    # 压掉探测不存在索引时的 DSHOW/MSMF WARN 刷屏
    try:
        cv2.utils.logging.setLogLevel(cv2.utils.logging.LOG_LEVEL_ERROR)
    except Exception:
        pass

    # DSHOW 失败换 MSMF 后端, 应对驱动不稳定
    cap = None
    for backend in (cv2.CAP_DSHOW, cv2.CAP_MSMF):
        cap = cv2.VideoCapture(device, backend)
        if cap.isOpened():
            break
        cap.release()
        cap = None
    if cap is None:
        return []

    # 外置 USB 摄像头默认 YUYV 高分辨率, USB 带宽大 → read 容易失败/黑屏。
    # 尽量切 MJPG 编码 + 小缓冲, 设置失败不影响使用。
    try:
        cap.set(cv2.CAP_PROP_FOURCC, cv2.VideoWriter_fourcc(*"MJPG"))
    except Exception:
        pass
    try:
        cap.set(cv2.CAP_PROP_BUFFERSIZE, 1)
    except Exception:
        pass
    res_w, res_h = _pick_resolution(cap)      # 自动取最高实际分辨率
    res_text = f"{res_w}x{res_h}"

    cache_dir = os.path.join(data_dir, "cache")
    os.makedirs(cache_dir, exist_ok=True)
    photos = []           # [(path, bytes, lum)]
    start = time.time()
    no_frame = 0          # 连续丢帧计数
    flash_pending = False # 拍照后下一帧闪白反馈
    warn_pending = ""     # 拍照亮度异常警告文本
    warn_frames = 0       # 警告剩余帧数 (~1.3s 自动消失)
    cv2.namedWindow(_WIN_TITLE, cv2.WINDOW_NORMAL)   # 可拖动缩放窗口
    cv2.resizeWindow(_WIN_TITLE, 960, int(960 * res_h / res_w))
    try:
        while time.time() - start < timeout_s:
            ret, frame = cap.read()
            if not ret:
                # 丢帧重试: 外置 USB 摄像头驱动起步慢/偶发丢帧,
                # 立即退出会表现为"摄像头一闪就关"。
                no_frame += 1
                if no_frame == 1:
                    blank = np.zeros((540, 960, 3), np.uint8)
                    cv2.putText(blank, "Waiting for camera frame...", (12, 30),
                                cv2.FONT_HERSHEY_SIMPLEX, 0.65, (0, 255, 0), 2)
                    cv2.imshow(_WIN_TITLE, blank)
                if no_frame >= 60:   # 连续约 2s 无帧才放弃
                    break
                cv2.waitKey(30)
                continue
            no_frame = 0
            # 实时亮度 (感知灰度均值), 供 HUD 显示 + 拍照异常提醒
            lum = _lum_of(frame)
            fh, fw = frame.shape[:2]
            scale = 960.0 / fw
            disp = cv2.resize(frame, (960, int(fh * scale)))
            warn_text = warn_pending if warn_frames > 0 else ""
            disp = _draw_hud(disp, len(photos), device, device_name, res_text,
                             lum=lum, warn_text=warn_text, flash=flash_pending)
            flash_pending = False
            if warn_frames > 0:
                warn_frames -= 1
            cv2.imshow(_WIN_TITLE, disp)
            k = cv2.waitKey(30) & 0xFF
            # 窗口被用户手动关闭 (点 X) → 视为取消, 干净退出 (防止 imshow 已销毁窗口后崩溃)
            if cv2.getWindowProperty(_WIN_TITLE, cv2.WND_PROP_VISIBLE) < 1:
                photos.clear()
                break
            if k in (13, 10):          # 回车: 结束
                break
            elif k == 32:              # 空格: 拍照 (存原帧, 分辨率不缩)
                ok, buf = cv2.imencode(".jpg", frame, [cv2.IMWRITE_JPEG_QUALITY, 90])
                if ok:
                    fname = f"cam_{int(time.time()*1000)}_{len(photos)}.jpg"
                    path = os.path.join(cache_dir, fname)
                    with open(path, "wb") as f:
                        f.write(buf.tobytes())
                    photos.append((path, buf.tobytes(), lum))
                    flash_pending = True
                    # 亮度异常: 浮层警告约 1.3s, 提醒重拍
                    if lum < LUM_DARK:
                        warn_pending, warn_frames = "!! LOW LIGHT !!", 40
                    elif lum > LUM_BRIGHT:
                        warn_pending, warn_frames = "!! TOO BRIGHT !!", 40
                if len(photos) >= max_photos:
                    break
            elif k == 8:               # Backspace: 撤回最近一张
                if photos:
                    path, _, _ = photos.pop()
                    try:
                        os.remove(path)
                    except OSError:
                        pass
            elif k == 27:              # ESC: 取消
                photos.clear()
                break
    finally:
        # 稍等驱动收尾再释放, 降低反复开关 USB 摄像头时驱动状态错乱概率
        time.sleep(0.3)
        try:
            cap.release()
        except Exception:
            pass
        try:
            cv2.destroyAllWindows()
        except Exception:
            pass
    return photos


# 探测结果缓存: 全量扫描要 5~10 秒, 30 秒内复用, 点"重新检测"强制重扫
_camera_cache = {"ts": 0.0, "devices": None}
_CAMERA_CACHE_TTL = 30.0


def _probe_index(idx: int) -> bool:
    """单个索引探测 (独立线程执行以便超时保护, 防 DSHOW 偶发卡死)。"""
    import time
    import cv2
    try:
        cv2.utils.logging.setLogLevel(cv2.utils.logging.LOG_LEVEL_ERROR)
    except Exception:
        pass
    for attempt in range(2):
        for backend in (cv2.CAP_DSHOW, cv2.CAP_MSMF):
            cap = cv2.VideoCapture(idx, backend)
            if not cap.isOpened():
                cap.release()
                continue
            # read 等待首帧, 重试 3 次
            for _ in range(3):
                ret, frame = cap.read()
                if ret and frame is not None:
                    cap.release()
                    return True
            cap.release()
        time.sleep(0.3)   # 驱动未就绪, 稍等重试
    return False


@app.route("/api/camera/list", methods=["GET"])
def camera_list():
    """扫描可用摄像头, 返回 [{index, ok, name}, ...] (探测 0~8)。

    每个设备: DSHOW 失败换 MSMF, 打开后 read 重试, 失败稍等再试一次,
    应对驱动初始化慢 / 摄像头刚插上的情况; 单个索引超时 4s 判失败,
    避免 DSHOW 偶发卡死拖死整个探测。
    name 来自 Windows PnP 设备名, 按探测到的顺序对应 (拿不到名称为空串)。
    ?force=1 强制重新扫描 (前端"🔄 重新检测")。
    """
    import time
    from concurrent.futures import ThreadPoolExecutor

    global _camera_cache
    if _camera_cache["devices"] is not None and \
            time.time() - _camera_cache["ts"] < _CAMERA_CACHE_TTL and \
            request.args.get("force") != "1":
        return jsonify({"ok": True, "devices": _camera_cache["devices"]})

    names = _camera_names()
    ok_flags = {}
    with ThreadPoolExecutor(max_workers=9) as ex:
        futs = {ex.submit(_probe_index, i): i for i in range(9)}
        for fut, idx in futs.items():
            try:
                ok_flags[idx] = fut.result(timeout=4)
            except Exception:
                ok_flags[idx] = False
    devices = []
    ok_order = 0
    for idx in range(9):
        ok = bool(ok_flags.get(idx))
        name = names[ok_order] if ok and ok_order < len(names) else ""
        if ok:
            ok_order += 1
        devices.append({"index": idx, "ok": ok, "name": name})
    _camera_cache = {"ts": time.time(), "devices": devices}
    return jsonify({"ok": True, "devices": devices})


@app.route("/api/camera/capture", methods=["POST"])
def camera_capture():
    """打开指定摄像头拍照 (空格拍照/回车结束), 逐张 OCR, 照片识别后删除。

    body: {device: 摄像头编号}  (默认 1)
    返回 {ok, n, groups: [[行,...], ...]} (每张图一组文本行)。
    """
    body = request.get_json(force=True, silent=True) or {}
    device = int(body.get("device", 1))
    data_dir = get_data_dir()
    # 从探测缓存里带出设备名 (HUD 顶部显示用, 查不到就用 Cam N)
    dev_name = ""
    if _camera_cache["devices"]:
        for d in _camera_cache["devices"]:
            if d.get("index") == device:
                dev_name = d.get("name") or ""
                break
    photos = _camera_capture_session(data_dir, device=device, device_name=dev_name)
    if not photos:
        return jsonify({"error": f"摄像头 {device} 不可用或未拍到照片"}), 400
    try:
        from warehouse.ocr import recognize
        groups = []
        n_dark = 0
        n_bright = 0
        for path, buf, lum in photos:
            if lum < LUM_DARK:
                n_dark += 1
            elif lum > LUM_BRIGHT:
                n_bright += 1
            try:
                groups.append(recognize(buf))
            except Exception as e:
                groups.append([])
            finally:
                # 照片用完即删 (缓存不残留)
                try:
                    if os.path.exists(path):
                        os.remove(path)
                except OSError:
                    pass
        # 亮度异常提醒 (OCR 过暗/过亮会识别不准)
        warn_light = ""
        if n_dark and n_bright:
            warn_light = f"有 {n_dark} 张过暗、{n_bright} 张过亮，识别可能不准，建议补拍"
        elif n_dark:
            warn_light = f"有 {n_dark} 张照片环境过暗，识别可能不准，建议开灯补拍"
        elif n_bright:
            warn_light = f"有 {n_bright} 张照片过亮/反光，识别可能不准，建议避开强光补拍"
        return jsonify({"ok": True, "n": len(photos), "groups": groups,
                        "warn_light": warn_light})
    except Exception as e:
        return jsonify({"error": f"识别失败: {str(e)[:200]}"}), 500


# ── API: 数据包 导出/导入 (ZIP) ────────────────────────
@app.route("/api/data/export", methods=["POST"])
def data_export():
    """一键导出全部元器件数据为 ZIP。

    body: {out_dir: 导出目录, 缺省存 data/exports/}
    """
    from warehouse.packfile import export_package
    body = request.get_json(force=True, silent=True) or {}
    try:
        r = export_package(get_data_dir(), out_dir=body.get("out_dir") or "")
        return jsonify(r)
    except Exception as e:
        return jsonify({"error": f"导出失败: {str(e)[:200]}"}), 500


@app.route("/api/data/import", methods=["POST"])
def data_import():
    """一键导入 ZIP 数据包 (合并模式, 导入前自动备份)。"""
    from warehouse.packfile import import_package
    if "file" not in request.files:
        return jsonify({"error": "未收到数据包文件"}), 400
    f = request.files["file"]
    if not (f.filename or "").lower().endswith(".zip"):
        return jsonify({"error": "仅支持 .zip 数据包"}), 400
    try:
        r = import_package(get_data_dir(), f.read(), project_root=BASE_DIR)
        return jsonify(r)
    except Exception as e:
        return jsonify({"error": f"导入失败: {str(e)[:200]}"}), 500


# ── API: 批量导入 ──────────────────────────────────────
@app.route("/api/import_parse_text", methods=["POST"])
def import_parse_text():
    data = request.get_json(force=True)
    text = (data.get("text") or "").strip()
    if not text:
        return jsonify({"error": "内容为空"}), 400
    cfg = get_ai_cfg()
    if not cfg:
        return jsonify({"error": "未配置 AI 接口，请到「设置 → AI」填写 API Key"}), 400
    try:
        parser = BatchParser(cfg["api_key"], cfg["base_url"], cfg["model"])
        items = parser.parse_text(text)
        return jsonify({"ok": True, "items": items, "dropped_nc": parser.dropped_nc,
                        "usage": parser.usage})
    except Exception as e:
        return jsonify({"error": str(e)[:300]}), 500


@app.route("/api/import_parse_excel", methods=["POST"])
def import_parse_excel():
    cfg = get_ai_cfg()
    if not cfg:
        return jsonify({"error": "未配置 AI 接口，请到「设置 → AI」填写 API Key"}), 400
    if "file" not in request.files:
        return jsonify({"error": "未收到文件"}), 400
    f = request.files["file"]
    fname = f.filename or ""
    if not fname.lower().endswith((".xlsx", ".xls")):
        return jsonify({"error": "仅支持 .xlsx / .xls 文件"}), 400
    raw_file = f.read()
    if not raw_file:
        return jsonify({"error": "导入文件为空"}), 400
    try:
        parser = BatchParser(cfg["api_key"], cfg["base_url"], cfg["model"])
        items, preview = parser.parse_excel(raw_file, fname)
        return jsonify({"ok": True, "items": items, "preview": preview,
                        "filename": fname, "dropped_nc": parser.dropped_nc,
                        "usage": parser.usage})
    except Exception as e:
        return jsonify({"error": str(e)[:300]}), 500


@app.route("/api/import_parse_rules", methods=["POST"])
def import_parse_rules():
    """脚本解析 (纯规则, 无 AI): 支持粘贴文本 {text} 或上传文件 (excel/txt)。

    返回 {ok, items, dropped_nc, mode: "rules"}。
    """
    from warehouse.rules import RuleParser
    text = ""
    if "file" in request.files:
        f = request.files["file"]
        fname = (f.filename or "").lower()
        data = f.read()
        if fname.endswith((".xlsx", ".xls")):
            from io import BytesIO
            from openpyxl import load_workbook
            wb = load_workbook(BytesIO(data), data_only=True)
            ws = wb.active
            lines = []
            for row in ws.iter_rows(values_only=True):
                cells = [str(c).strip() for c in row if c is not None and str(c).strip()]
                if cells:
                    lines.append(" ".join(cells))
            wb.close()
            text = "\n".join(lines)
        else:
            try:
                text = data.decode("utf-8")
            except UnicodeDecodeError:
                text = data.decode("gbk", errors="ignore")
        if not text.strip():
            return jsonify({"error": "文件内容为空"}), 400
    else:
        body = request.get_json(force=True, silent=True) or {}
        text = (body.get("text") or "").strip()
        if not text:
            return jsonify({"error": "内容为空"}), 400
    try:
        parser = RuleParser()
        items = parser.parse_text(text)
        return jsonify({"ok": True, "items": items, "dropped_nc": parser.dropped_nc,
                        "mode": "rules"})
    except Exception as e:
        return jsonify({"error": str(e)[:300]}), 500


@app.route("/api/import_commit", methods=["POST"])
def import_commit():
    data = request.get_json(force=True)
    items = data.get("items", [])
    if not items:
        return jsonify({"error": "没有可写入的数据"}), 400
    cfg = get_ai_cfg()
    if not cfg:
        return jsonify({"error": "未配置 AI 接口"}), 400
    try:
        parser = BatchParser(cfg["api_key"], cfg["base_url"], cfg["model"])
        result = parser.commit(items, get_data_dir())
        return jsonify({"ok": True, "result": result, "total": len(items)})
    except Exception as e:
        return jsonify({"error": str(e)[:300]}), 500


# ── API: AI 快速填入 ───────────────────────────────────
@app.route("/api/ai_fill", methods=["POST"])
def ai_fill():
    data = request.get_json(force=True)
    name = (data.get("subcat") or "").strip()
    desc = (data.get("desc") or "").strip()
    if not name or primary_owner(name) is None:
        return jsonify({"error": "未知子分类"}), 404
    if not desc:
        return jsonify({"error": "描述为空"}), 400
    cfg = get_ai_cfg()
    if not cfg:
        return jsonify({"error": "未配置 AI 接口，请到「设置 → AI」填写 API Key"}), 500
    try:
        fields = fields_for(name)
        parsed = AIFiller(
            api_key=cfg["api_key"], base_url=cfg["base_url"], model=cfg["model"]
        ).parse(name, desc, fields, subcat_list=None)
        return jsonify({"ok": True, "item": parsed})
    except Exception as e:
        return jsonify({"error": str(e)}), 500


def main():
    threading.Timer(1.2, lambda: webbrowser.open("http://127.0.0.1:5000")).start()
    # threaded=True: 摄像头拍照/探测等阻塞端点不能卡死整个服务
    app.run(host="127.0.0.1", port=5000, debug=False, threaded=True)


if __name__ == "__main__":
    main()
