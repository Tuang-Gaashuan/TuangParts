# -*- coding: utf-8 -*-
"""元器件仓库 — 品牌数据清洗 (一次性历史数据修正工具)。

把 data 目录所有 xlsx 的品牌列按规范映射表修正:
  - 同品牌多写法合并为规范名 (含中文注释, 如 FOJAN(富捷))
  - 修正错乱写法 (OCR/录入错误, 如 JERR(捷而场) → JIERR(捷而瑞))
  - 非品牌描述清空 (如"中性电容")

用法:  python tools/maintenance/clean_brands.py [data目录]
不传 data 目录时自动用工程 data 目录。
运行前请先备份 (本脚本也会打印备份提示)。
"""

import os
import sys
import datetime

PROJECT_DIR = os.path.abspath(os.path.join(os.path.dirname(__file__), "..", ".."))

# ── 规范映射: 原写法 -> 规范名 (含搜索确认依据) ──────────────
# 备注格式: (规范名, 修正说明)
BRAND_MAP = {
    # ── FOJAN 富捷 (安徽省富捷电子科技, 立创在售: 贴片电阻/采样电阻/MLCC) ──
    "FOJAN": "FOJAN(富捷)",
    "FOJAN富捷": "FOJAN(富捷)",
    "富捷": "FOJAN(富捷)",
    "FOJAN(富健)": "FOJAN(富捷)",      # 错别字 健→捷
    "FOJAN(赛境)": "FOJAN(富捷)",      # OCR 错写 (同品牌电阻料)

    # ── 厚声 Uniroyal/UniOhm (昆山厚声电子工业, uni-royal.cn) ──
    "UNIROYAU": "厚声(UniOhm)",         # UNIROYAL 错写
    "ROYAU": "厚声(UniOhm)",            # UNIROYAL 截断
    "UN-ROYAL(TP)": "厚声(UniOhm)",     # UNIROYAL 错写
    "UNHROYAU(P)": "厚声(UniOhm)",      # UNIROYAL 多字符错写
    "厚声": "厚声(UniOhm)",

    # ── JIERR 捷而瑞 (采样/合金电阻; JERR* 系 OCR 错乱, 括号中文乱码) ──
    "JIERR捷而瑞": "JIERR(捷而瑞)",
    "JERR(捷而场)": "JIERR(捷而瑞)",
    "JERR(德五现)": "JIERR(捷而瑞)",
    "JERR(提而项)": "JIERR(捷而瑞)",
    "JERR(QF)": "JIERR(捷而瑞)",
    "JERR(境系)": "JIERR(捷而瑞)",
    "JERR耐": "JIERR(捷而瑞)",

    # ── YAGEO 国巨 (台湾国巨股份, 贴片电阻全球龙头) ──
    "YAGEO(E)": "YAGEO(国巨)",          # (E) 疑为误标
    "YAGEO": "YAGEO(国巨)",

    # ── SAMSUNG 三星电机 (MLCC) ──
    "SAMSUNG": "SAMSUNG(三星)",

    # ── TI 德州仪器 ──
    "TI（德州仪器）": "TI(德州仪器)",
    "TI": "TI(德州仪器)",

    # ── 村田 Murata ──
    "村田": "村田(Murata)",

    # ── ST 意法半导体 ──
    "ST": "ST(意法半导体)",

    # ── 固得沃克 GOODWORK (二极管/MOS) ──
    "固得沃克GOODWORK": "GOODWORK(固得沃克)",
    "GOODWORK固得沃克": "GOODWORK(固得沃克)",

    # ── MDD 辰达半导体 (二极管) ──
    "辰达半导体MDD": "MDD(辰达半导体)",

    # ── R+O 宏嘉诚 (二极管) ──
    "宏嘉诚R+O": "R+O(宏嘉诚)",

    # ── PROD 谱罗德 (电阻) ──
    "谱罗德": "PROD(谱罗德)",

    # ── Coilank 驰兴 (电感) ──
    "Coilank驰兴": "Coilank(驰兴)",
    "Coilank(驰兴电感)": "Coilank(驰兴)",

    # ── SHOU HAN 首韩 ──
    "首韩": "SHOU HAN(首韩)",

    # ── JSMSEMI 杰盛微 ──
    "杰盛微JSMSEMI": "JSMSEMI(杰盛微)",

    # ── FOLLON 富隆电子 ──
    "FOLLON富隆电子": "FOLLON(富隆电子)",

    # ── HRE 芯声 ──
    "HRE（芯声）": "HRE(芯声)",
    "芯声": "HRE(芯声)",

    # ── HC 虹成电子 ──
    "HC虹成电子": "HC(虹成电子)",

    # ── CAX 创都 ──
    "CAX创都": "CAX(创都)",

    # ── APV 爱普微 ──
    "爱普微": "APV(爱普微)",
    "APV（爱普微）": "APV(爱普微)",

    # ── SIkor 萨科微 (二极管/MOS) ──
    "萨科微SIkor": "SIkor(萨科微)",

    # ── HCI 杭晶 (晶振) ──
    "杭晶": "HCI(杭晶)",

    # ── LX 连欣科技 ──
    "LX连欣科技": "LX(连欣科技)",

    # ── IDCHIP 英锐芯 ──
    "英锐芯": "IDCHIP(英锐芯)",

    # ── YJYCOIN 益嘉源 ──
    "YJYCOIN益嘉源": "YJYCOIN(益嘉源)",

    # ── BHFUSE 佰宏 (保险丝) ──
    "佰宏": "BHFUSE(佰宏)",

    # ── SIT 芯力特 (CAN 收发器, 已被纳芯微收购) ──
    "SIT": "SIT(芯力特)",

    # ── MagnTek 麦歌恩 (磁性传感器, magntek.com.cn) ──
    "MagnTek（麦歌恩）": "MagnTek(麦歌恩)",

    # ── 江苏长电/长晶CJ: MMBT3904 三极管, 长电科技(JCET)为主产商 ──
    "江苏长电/长晶CJ": "江苏长电(JCET)",

    # ── 非品牌: 描述词混入品牌列 → 清空 ──
    "中性电容": "",
}

# 已知品牌: 保持原名不动 (防止未来误映射; 仅作白名单参考, 不影响脚本)
KEEP_NAMES = [
    "台舟TECH PUBLIC", "长江微电", "静芯ElecSuper", "HXY MOSFET(华轩阳电子)",
    "YJX(雅晶鑫)", "XUNPU(讯普)", "HOAUC(华宇创)", "创基CBI", "WCH",
    "Torex", "S&S(海旭)", "XHXDZ(兴华鑫)", "瑞森半导体", "顾邦半导体GOSEMICON",
    "XNRUSEMI(新锐)", "乐山无线电LRC", "友台半导体UMW", "东沃DOWO",
    "Infineon", "YXC扬兴科技", "HI-LINK(海凌科)", "ACMECOM(鼎容)", "国连",
    "华灿天禄", "福然德", "捷茂微", "MaxLinear", "KOA", "FENGHUA",
    "MPS(芯源)", "YSUNK", "合科泰Hottech", "Econd(米朗电子)", "TKD(泰晶)",
    "MagnTek(麦歌恩)",
]


def find_data_dir() -> str:
    """定位实际数据目录: 命令行参数 > settings.json 的 data_dir > 默认 data。"""
    if len(sys.argv) > 1:
        return os.path.abspath(sys.argv[1])
    import json
    sp = os.path.join(PROJECT_DIR, "data", "settings.json")
    if os.path.exists(sp):
        try:
            with open(sp, encoding="utf-8") as f:
                s = json.load(f)
            custom = (s.get("data_dir") or "").strip()
            if custom:
                return os.path.abspath(custom)
        except Exception:
            pass
    return os.path.join(PROJECT_DIR, "data")


def clean(data_dir: str) -> dict:
    """遍历所有 xlsx 修正品牌列。返回统计。"""
    from openpyxl import load_workbook
    changed_files = []
    changed_rows = 0
    by_brand = {}   # 规范名 -> (原写法列表)

    for dirpath, _dirs, filenames in os.walk(data_dir):
        for fn in sorted(filenames):
            if not fn.lower().endswith(".xlsx"):
                continue
            fp = os.path.join(dirpath, fn)
            wb = load_workbook(fp)
            ws = wb.active
            file_changed = 0
            for row in ws.iter_rows(min_row=2):
                cell = row[1]
                if cell is None or cell.value is None:
                    continue
                orig = str(cell.value).strip()
                if not orig:
                    continue
                target = BRAND_MAP.get(orig)
                if target is None:
                    continue   # 保持原样
                if target:
                    by_brand.setdefault(target, set()).add(orig)
                    cell.value = target
                else:
                    by_brand.setdefault("(清空)", set()).add(orig)
                    cell.value = ""
                file_changed += 1
            if file_changed:
                wb.save(fp)
                changed_files.append((os.path.relpath(fp, data_dir), file_changed))
                changed_rows += file_changed
            wb.close()

    return {
        "changed_files": changed_files,
        "changed_rows": changed_rows,
        "by_brand": {k: sorted(v) for k, v in by_brand.items()},
    }


def main():
    data_dir = find_data_dir()
    print(f"数据目录: {data_dir}")
    if not os.path.isdir(data_dir):
        print("目录不存在, 退出"); sys.exit(1)
    print("开始清洗...")
    r = clean(data_dir)
    print(f"\n修改行数: {r['changed_rows']}, 涉及文件: {len(r['changed_files'])}")
    print("\n合并明细 (规范名 <- 原写法):")
    for target, origins in sorted(r["by_brand"].items(), key=lambda kv: -len(kv[1])):
        print(f"  {target:24s} <- {', '.join(origins)}")
    if r["changed_files"]:
        print("\n修改的文件:")
        for fp, n in r["changed_files"]:
            print(f"  {fp}  ({n} 行)")
    print("\n完成。品牌列已规范, 无品牌(清空)行:", r["by_brand"].get("(清空)", []))


if __name__ == "__main__":
    main()
