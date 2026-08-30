# -*- coding: utf-8 -*-
"""元器件仓库 — 数据包导出/导入 (ZIP)。

导出: 所有分类 xlsx + 未分类 + manifest.json → data/exports/parts-warehouse_导出_<时间戳>.zip
      排除运行时数据 (cache/backgrounds/decor/日志/设置/旧导出包)。
导入: 合并模式, 同名子分类按四键规则合并数量, 导入前自动备份到 backups/。
"""
import datetime
import io
import json
import os
import re
import shutil
import zipfile

from warehouse.config import COMMON_FIELDS

# 导出一级分类目录时排除的运行时目录/文件
RUNTIME_NAMES = {"cache", "backgrounds", "decor", "exports", "backups",
                 "settings.json", "activity_log.jsonl", "ledger.jsonl", "undo_log.jsonl",
                 "撤回日志", "未分类", "BOM清单"}

EXPORT_SUBDIR = "exports"
BACKUP_ROOT_REL = "backups"     # 项目根下的 backups/ (与 pack.py 一致)


def _category_dirs(data_dir: str) -> list:
    """data 下一级分类目录 (排除运行时目录)。"""
    dirs = []
    for name in sorted(os.listdir(data_dir)):
        p = os.path.join(data_dir, name)
        if os.path.isdir(p) and name not in RUNTIME_NAMES:
            dirs.append(p)
    return dirs


def _count_items(xlsx_path: str) -> int:
    """统计 xlsx 数据行数。"""
    try:
        from openpyxl import load_workbook
        wb = load_workbook(xlsx_path, read_only=True, data_only=True)
        ws = wb.active
        n = 0
        for row in ws.iter_rows(min_row=2, values_only=True):
            if any(c is not None and str(c).strip() for c in row):
                n += 1
        wb.close()
        return n
    except Exception:
        return 0


def export_package(data_dir: str, out_dir: str = "") -> dict:
    """打包全部元器件数据为 ZIP。

    out_dir 为空 → data/exports/; 否则导出到指定目录。
    返回 {ok, path, filename, subcats, items}。
    """
    from warehouse import unclassified
    export_dir = out_dir.strip() or os.path.join(data_dir, EXPORT_SUBDIR)
    os.makedirs(export_dir, exist_ok=True)
    ts = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"parts-warehouse_导出_{ts}.zip"
    out_path = os.path.join(export_dir, filename)

    subcats = 0
    items = 0
    buf = io.BytesIO()
    with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as zf:
        for cat_dir in _category_dirs(data_dir):
            cat_name = os.path.basename(cat_dir)
            for fname in sorted(os.listdir(cat_dir)):
                if not fname.lower().endswith(".xlsx"):
                    continue
                fp = os.path.join(cat_dir, fname)
                n = _count_items(fp)
                if n == 0:
                    continue          # 空分类不打包
                zf.write(fp, os.path.join(cat_name, fname))
                subcats += 1
                items += n
        # 未分类
        h_u, rows = unclassified.load(data_dir)
        if rows:
            uncat_dir = os.path.join(data_dir, "未分类")
            os.makedirs(uncat_dir, exist_ok=True)
            uncat_path = os.path.join(uncat_dir, "未分类.xlsx")
            from openpyxl import Workbook
            wb = Workbook()
            ws = wb.active
            ws.append([label for _, label in COMMON_FIELDS])
            for r in rows:
                ws.append(r)
            wb.save(uncat_path)
            zf.write(uncat_path, "未分类/未分类.xlsx")
            subcats += 1
            items += len(rows)
        # BOM 清单：独立业务数据，原样打包，不能作为库存分类 Excel 处理。
        from warehouse import bom_lists
        bom_dir = os.path.join(data_dir, bom_lists.DIR_NAME)
        bom_lists_exported = 0
        if os.path.isdir(bom_dir):
            for fname in sorted(os.listdir(bom_dir)):
                if fname.lower().endswith(".xlsx"):
                    zf.write(os.path.join(bom_dir, fname), os.path.join(bom_lists.DIR_NAME, fname))
                    bom_lists_exported += 1
        # manifest
        manifest = {
            "app": "元器件仓库", "format_version": 2,
            "exported_at": ts, "subcats": subcats, "items": items,
            "bom_lists": bom_lists_exported,
        }
        zf.writestr("manifest.json", json.dumps(manifest, ensure_ascii=False, indent=2))

    with open(out_path, "wb") as f:
        f.write(buf.getvalue())
    return {"ok": True, "path": out_path, "filename": filename,
            "subcats": subcats, "items": items}


def _extract_safe(zf: zipfile.ZipFile, tmp_dir: str):
    """安全解压 (防路径穿越)。返回 xlsx 文件相对路径列表。"""
    files = []
    for info in zf.infolist():
        name = info.filename.replace("\\", "/")
        if name.startswith("/") or ".." in name.split("/"):
            raise ValueError(f"数据包包含非法路径: {name}")
        target = os.path.join(tmp_dir, name)
        if not os.path.abspath(target).startswith(os.path.abspath(tmp_dir)):
            raise ValueError(f"数据包路径越界: {name}")
        if info.is_dir():
            os.makedirs(target, exist_ok=True)
            continue
        os.makedirs(os.path.dirname(target), exist_ok=True)
        with zf.open(info) as src, open(target, "wb") as dst:
            shutil.copyfileobj(src, dst)
        if name.lower().endswith(".xlsx"):
            files.append(name)
    return files


def _norm_subcat(name: str) -> str:
    """子分类名归一化: 去空格/连字符, 斜杠去除, 小写。用于旧名→标准名匹配。"""
    return re.sub(r"[\s\-—/]+", "", name or "").lower()


def _find_subcat(name: str) -> str | None:
    """在分类体系里找子分类: 精确 > 归一化模糊。"""
    from warehouse.config import CATEGORIES
    for cats in CATEGORIES.values():
        for sc in cats[1]:
            if sc == name:
                return sc
    n = _norm_subcat(name)
    for cats in CATEGORIES.values():
        for sc in cats[1]:
            if _norm_subcat(sc) == n:
                return sc
    return None


def _find_subcat_in_category(category_name: str, filename: str) -> str | None:
    """按数据包保存的一级分类目录和文件名恢复标准子分类。"""
    from warehouse.config import CATEGORIES

    stem = os.path.splitext(os.path.basename(filename))[0]
    category_norm = _norm_subcat(category_name)
    candidates = []
    for _key, (cat_name, subs) in CATEGORIES.items():
        if _norm_subcat(cat_name) == category_norm:
            candidates.extend(subs)
    for candidate in candidates:
        if candidate == stem:
            return candidate
    stem_norm = _norm_subcat(stem)
    for candidate in candidates:
        if _norm_subcat(candidate) == stem_norm:
            return candidate
    return None


def import_package(data_dir: str, zip_bytes: bytes, project_root: str = "") -> dict:
    """合并导入 ZIP 数据包。

    返回 {ok, subcats, items, backup, detail: {子分类: 导入条数}}。
    """
    import tempfile
    from openpyxl import load_workbook
    from warehouse.excel_store import ExcelStore
    from warehouse.batch_import import _merge_rows
    from warehouse.activity import record as record_activity
    from warehouse import unclassified
    from warehouse import undo as undo_mod

    # 1. 解压校验
    with zipfile.ZipFile(io.BytesIO(zip_bytes)) as zf:
        names = zf.namelist()
        if "manifest.json" not in names:
            # 兼容无 manifest 的包: 只要有 xlsx 就接受
            has_xlsx = any(n.lower().endswith(".xlsx") for n in names)
            if not has_xlsx:
                raise ValueError("不是有效的数据包: 未找到 manifest.json 或 xlsx 文件")
        with tempfile.TemporaryDirectory() as tmp:
            files = _extract_safe(zf, tmp)

            # 2. 导入前备份
            backup = ""
            if project_root:
                bk_root = os.path.join(project_root, BACKUP_ROOT_REL)
                os.makedirs(bk_root, exist_ok=True)
                ts = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
                backup = os.path.join(bk_root, f"parts-warehouse_data_{ts}")
                shutil.copytree(data_dir, backup)

            store = ExcelStore(data_dir)
            from warehouse import bom_lists
            existing_bom_ids = {item.get("id") for item in bom_lists._all_xlsx(data_dir)}
            imported_bom_lists = 0
            skipped_bom_lists = 0
            detail = {}
            total_items = 0
            for rel in files:
                fp = os.path.join(tmp, rel)
                parts = rel.split("/")
                fname = os.path.basename(rel)
                # BOM 清单是独立的业务 Excel，绝不能按库存表合并或送入未分类。
                if len(parts) >= 2 and parts[0] == bom_lists.DIR_NAME:
                    bom_item = bom_lists._read_xlsx(fp)
                    if not bom_item or not bom_item.get("id"):
                        detail[f"BOM清单/{fname}(跳过)"] = "不是有效的 BOM 清单"
                        continue
                    if bom_item["id"] in existing_bom_ids:
                        skipped_bom_lists += 1
                        continue
                    target_dir = os.path.join(data_dir, bom_lists.DIR_NAME)
                    os.makedirs(target_dir, exist_ok=True)
                    target = os.path.join(target_dir, fname)
                    stem, ext = os.path.splitext(fname)
                    suffix = 2
                    while os.path.exists(target):
                        target = os.path.join(target_dir, f"{stem}_{suffix}{ext}")
                        suffix += 1
                    shutil.copy2(fp, target)
                    existing_bom_ids.add(bom_item["id"])
                    imported_bom_lists += 1
                    continue
                # 撤回日志是操作快照 (undo-*.xlsx), 不是库存数据。
                # 旧版导出包可能包含该目录, 导入时必须整目录跳过,
                # 否则快照列 (原分类/原行号/...) 会被错位当成库存行。
                if fname.startswith("undo-") or (len(parts) >= 2 and parts[0] == "撤回日志"):
                    continue
                # 未分类
                if fname == "未分类.xlsx" or (len(parts) >= 2 and parts[0] == "未分类"):
                    h_u, old_uncat = unclassified.load(data_dir)
                    from openpyxl import load_workbook
                    wb = load_workbook(fp, data_only=True)
                    ws = wb.active
                    grp = []
                    for row in ws.iter_rows(min_row=2, values_only=True):
                        if any(c is not None and str(c).strip() for c in row):
                            grp.append(list(row))
                    wb.close()
                    if grp:
                        unclassified.add(data_dir, grp)
                        undo_mod.push(data_dir, [
                            {"subcat": undo_mod.UNCAT_KEY, "old_rows": old_uncat},
                        ], "导入数据包(未分类)")
                    detail["未分类"] = detail.get("未分类", 0) + len(grp)
                    total_items += len(grp)
                    continue
                # 常规子分类: 优先使用 ZIP 路径中的一级分类目录恢复。
                category_name = parts[-2] if len(parts) >= 2 else ""
                subcat = _find_subcat_in_category(category_name, fname)
                if subcat is None:
                    subcat = _find_subcat(os.path.splitext(fname)[0])
                    if subcat is None:
                        subcat = os.path.splitext(fname)[0]
                try:
                    old_headers, old_rows = store.load(subcat)
                except ValueError:
                    # 旧命名 (如 "TVS-ESD" vs 标准 "TVS / ESD"): 归一化匹配标准名
                    std = _find_subcat(subcat)
                    if std:
                        subcat = std
                        try:
                            old_headers, old_rows = store.load(subcat)
                        except ValueError:
                            old_headers, old_rows = [], []
                    else:
                        # 旧包的目录名/文件名可能已经变化，但导出行仍保留“子分类”。
                        # 逐行恢复可识别分类，只有没有合法分类信息的行才进未分类。
                        wb = load_workbook(fp, data_only=True)
                        ws = wb.active
                        recoverable = {}
                        unknown_rows = []
                        for row in ws.iter_rows(min_row=2, values_only=True):
                            if not any(c is not None and str(c).strip() for c in row):
                                continue
                            r = list(row)
                            r = (r + [""] * len(COMMON_FIELDS))[:len(COMMON_FIELDS)]
                            row_subcat = _find_subcat(str(r[5] or ""))
                            if row_subcat:
                                recoverable.setdefault(row_subcat, []).append(r)
                            else:
                                unknown_rows.append({
                                    "name": str(r[0] or ""), "brand": str(r[1] or ""),
                                    "package": str(r[2] or ""), "qty": str(r[3] or "10"),
                                    "spec": str(r[6] or ""),
                                    "raw": f"[原分类:{subcat}]",
                                })
                        wb.close()
                        headers = [label for _, label in COMMON_FIELDS]
                        for recovered_subcat, recovered_rows in recoverable.items():
                            old_headers, old_rows = store.load(recovered_subcat)
                            new_items = [{
                                "name": str(r[0] or ""), "brand": str(r[1] or ""),
                                "package": str(r[2] or ""), "qty": str(r[3] or "10"),
                                "spec": str(r[6] or ""),
                            } for r in recovered_rows]
                            merged = _merge_rows(old_rows, new_items, recovered_subcat)
                            path = store.save(recovered_subcat, headers, merged)
                            record_activity(data_dir, recovered_subcat, old_rows, merged, path=path)
                            undo_mod.push(data_dir, [{"subcat": recovered_subcat, "old_rows": old_rows}],
                                          "导入数据包")
                            added = len(merged) - len(old_rows)
                            detail[recovered_subcat] = detail.get(recovered_subcat, 0) + added
                            total_items += added
                        if unknown_rows:
                            h_u, old_uncat = unclassified.load(data_dir)
                            unclassified.add(data_dir, unknown_rows)
                            undo_mod.push(data_dir, [
                                {"subcat": undo_mod.UNCAT_KEY, "old_rows": old_uncat},
                            ], "导入数据包(未分类)")
                            detail[f"{subcat}→未分类"] = len(unknown_rows)
                            total_items += len(unknown_rows)
                        continue
                try:
                    headers = [label for _, label in COMMON_FIELDS]
                    wb = load_workbook(fp, data_only=True)
                    ws = wb.active
                    new_rows = []
                    for row in ws.iter_rows(min_row=2, values_only=True):
                        if any(c is not None and str(c).strip() for c in row):
                            new_rows.append(list(row))
                    wb.close()
                    if not new_rows:
                        continue
                    # 包内行(list) 转 dict, 供 _merge_rows 合并 (四键去重+数量累加)
                    new_items = [{
                        "name": str(r[0] or ""), "brand": str(r[1] or ""),
                        "package": str(r[2] or ""), "qty": str(r[3] or "10"),
                        "spec": str(r[6] if len(r) > 6 else ""),
                    } for r in new_rows]
                    merged = _merge_rows(old_rows, new_items, subcat)
                    path = store.save(subcat, headers, merged)
                    record_activity(data_dir, subcat, old_rows, merged, path=path)
                    undo_mod.push(data_dir, [{"subcat": subcat, "old_rows": old_rows}],
                                  "导入数据包")
                    added = len(merged) - len(old_rows)
                    detail[subcat] = added
                    total_items += added
                except Exception as e:
                    detail[f"{subcat}(跳过)"] = str(e)[:60]
            return {"ok": True, "subcats": len(detail), "items": total_items,
                    "bom_lists": imported_bom_lists, "bom_lists_skipped": skipped_bom_lists,
                    "backup": backup, "detail": detail}
