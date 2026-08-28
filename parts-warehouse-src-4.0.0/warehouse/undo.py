# -*- coding: utf-8 -*-
"""元器件仓库 — 撤回系统 (undo)。

每次修改库存的操作 (取出/增加/保存/合并/批量导入/归类/删除) 自动把
操作前的数据快照保存为「撤回日志」目录下的一个独立 Excel 文件：
    data/撤回日志/<undo_id>_<动作>_<时间戳>.xlsx
一个文件 = 一次操作，文件内容即该次操作涉及的物料清单（含原行号）。

撤回语义（用户方案）：
- 整笔撤回 = 把该 Excel 全部物料按录入写回库存，然后删除这个文件
- 单项撤回 = 只把指定行写回库存，并从 Excel 中删除该行（其余行保留，
  可继续单项撤回直到文件清空后自动删除）

旧的 undo_log.jsonl 记录保留只读兼容：新记录不再写 jsonl，但旧记录的
整笔/单项撤回仍可走旧路径。
"""
import json
import os
import uuid
from datetime import datetime

from openpyxl import Workbook, load_workbook

from warehouse.config import COMMON_FIELDS

UNDO_FILE = "undo_log.jsonl"
MAX_ENTRIES = 60   # 旧 jsonl 最多保留 60 条（仅兼容旧记录）
UNCAT_KEY = "__unclassified__"
SNAP_DIR = "撤回日志"
# 撤回日志 Excel 列: 子分类 / 原行号 / 通用字段
SNAP_COLS = ["子分类", "原行号"] + [label for _, label in COMMON_FIELDS]


def _path(data_dir: str) -> str:
    return os.path.join(data_dir, UNDO_FILE)


def _snap_dir(data_dir: str) -> str:
    return os.path.join(data_dir, SNAP_DIR)


# ── 撤回日志 Excel 读写 ───────────────────────────────────

def _snap_filename(undo_id: str, action: str, dt: datetime) -> str:
    safe_action = "".join(ch for ch in str(action or "操作") if ch not in '\\/:*?"<>|')
    return f"{undo_id}_{safe_action}_{dt.strftime('%Y%m%d_%H%M%S')}.xlsx"


def _iter_snap_files(data_dir: str):
    d = _snap_dir(data_dir)
    if not os.path.isdir(d):
        return
    for name in sorted(os.listdir(d)):
        if name.lower().endswith(".xlsx"):
            yield os.path.join(d, name), name


def _snap_file_by_undo(data_dir: str, undo_id: str):
    for p, name in _iter_snap_files(data_dir):
        if name.startswith(str(undo_id) + "_"):
            return p
    return None


def _read_snap(path: str) -> list:
    """读取撤回日志 Excel → [{subcat, orig_row, fields}]。"""
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    rows = []
    first = True
    for row in ws.iter_rows(values_only=True):
        if first:
            first = False
            continue
        if not row or row[0] is None:
            continue
        subcat = str(row[0] or "")
        try:
            orig = int(row[1])
        except (TypeError, ValueError):
            orig = -1
        fields = [("" if v is None else v) for v in row[2:2 + len(COMMON_FIELDS)]]
        rows.append({"subcat": subcat, "orig_row": orig, "fields": fields})
    wb.close()
    return rows


def _append_note(row: list, text: str) -> None:
    """在库存行备注列追加一次状态，不覆盖原备注。"""
    note_index = len(COMMON_FIELDS) - 1
    while len(row) <= note_index:
        row.append("")
    old = str(row[note_index] or "").strip()
    row[note_index] = f"{old}；{text}" if old else text


def _snapshot_row(data_dir: str, undo_id: str, subcat: str, row: int):
    path = _snap_file_by_undo(data_dir, undo_id)
    if not path:
        return None
    for item in _read_snap(path):
        if item["subcat"] == subcat and item["orig_row"] == row:
            return item
    return None


def apply_operation(data_dir: str, undo_id: str, details: list, mode: str = "undo") -> dict | None:
    """按操作快照执行撤回或取消撤回，保留快照文件和原始明细。"""
    if mode not in ("undo", "restore") or not _snap_file_by_undo(data_dir, undo_id):
        return None
    from warehouse.excel_store import ExcelStore
    from warehouse.config import fields_for
    store = ExcelStore(data_dir)
    by_subcat = {}
    for detail in details:
        subcat = str(detail.get("subcat", ""))
        try:
            row_index = int(detail.get("row", -1))
        except (TypeError, ValueError):
            continue
        if _snapshot_row(data_dir, undo_id, subcat, row_index) is not None:
            by_subcat.setdefault(subcat, []).append((row_index, detail))
    changed = []
    for subcat, entries in by_subcat.items():
        if subcat == UNCAT_KEY:
            from warehouse import unclassified
            headers, current = unclassified.load(data_dir)
        else:
            headers, current = store.load(subcat)
        for row_index, detail in entries:
            snap = _snapshot_row(data_dir, undo_id, subcat, row_index)
            if snap is None:
                continue
            snap_fields = snap.get("fields") or [""] * len(COMMON_FIELDS)
            snap_name = str((snap_fields or [""])[0] or "").strip()
            # 定位当前行: 优先按原行号; 行被删(取出归零/合并)时按名称匹配重建。
            target = None
            if 0 <= row_index < len(current):
                current_name = str((current[row_index] or [""])[0] or "").strip()
                if not snap_name or current_name == snap_name:
                    target = row_index
            if target is None:
                # 行号错位或行已删除: 按快照名称在当前行中查找
                for i, cand in enumerate(current):
                    cand_name = str((cand or [""])[0] or "").strip()
                    if cand_name == snap_name:
                        target = i
                        break
            if target is None and mode == "undo":
                # 行已被删除 (如取完最后一件后 0 数量行不落盘): 用快照重建。
                # 快照 fields 含操作前数量, 重建后由下方 new_qty 统一重算, 先置空。
                rebuilt = list(snap_fields)
                while len(rebuilt) <= 3:
                    rebuilt.append("")
                rebuilt[3] = "0"
                current.append(rebuilt)
                target = len(current) - 1
            if target is None:
                continue
            amount = abs(int(detail.get("delta", 0) or 0))
            if not amount:
                continue
            old_qty = int(float(str(current[target][3]).strip() or 0)) if len(current[target]) > 3 else 0
            direction = 1 if int(detail.get("delta", 0) or 0) < 0 else -1
            if mode == "restore":
                direction *= -1
            new_qty = old_qty + direction * amount
            if new_qty < 0:
                raise ValueError(f"「{detail.get('name', '')}」库存不足，无法取消撤回")
            current[target][3] = str(new_qty)
            label = "已撤回" if mode == "undo" else "取消撤回"
            _append_note(current[target], f"{label} {amount} 件")
            changed.append({"subcat": subcat, "row": target, "detail_index": detail.get("detail_index"), "name": detail.get("name", ""),
                            "delta": direction * amount, "quantity_before": old_qty,
                            "quantity_after": new_qty, "status": label})
        if subcat == UNCAT_KEY:
            unclassified._save(data_dir, headers, current)
        else:
            store.save(subcat, [label for _, label in fields_for(subcat)], current)
    return {"undo_id": undo_id, "mode": mode, "items": changed,
            "status": "已撤回" if mode == "undo" else "正常"}


def _write_snap(path: str, rows: list) -> None:
    os.makedirs(os.path.dirname(path), exist_ok=True)
    wb = Workbook()
    ws = wb.active
    ws.title = "快照"
    ws.append(SNAP_COLS)
    for r in rows:
        ws.append([r["subcat"], r["orig_row"]] + r["fields"])
    wb.save(path)


# ── 快照记录 (写入撤回日志 Excel) ───────────────────────── ─────────────────────────

def push(data_dir: str, snapshots: list, action: str = ""):
    """记录一次可撤回操作。snapshots: [{subcat, old_rows}, ...]。"""
    if not snapshots:
        return None
    undo_id = "undo-" + uuid.uuid4().hex[:12]
    dt = datetime.now()
    path = os.path.join(_snap_dir(data_dir), _snap_filename(undo_id, action, dt))
    rows = []
    for snap in snapshots:
        subcat = snap.get("subcat", "")
        old_rows = snap.get("old_rows") or []
        for i, row in enumerate(old_rows):
            vals = list(row) + [""] * (len(COMMON_FIELDS) - len(row))
            rows.append({"subcat": subcat, "orig_row": i, "fields": vals[:len(COMMON_FIELDS)]})
    _write_snap(path, rows)
    return {
        "time": dt.strftime("%Y-%m-%d %H:%M:%S"),
        "undo_id": undo_id,
        "action": action,
        "snapshots": snapshots,
        "_path": path,
    }


# ── 读取快照 (Excel + 旧 jsonl 合并) ──────────────────────

def _snap_info_from_filename(name: str) -> dict:
    parts = name.split("_")
    undo_id = parts[0] if parts else ""
    action = parts[1] if len(parts) > 1 else ""
    ts = "_".join(parts[2:]) if len(parts) > 2 else ""
    time_str = ""
    if ts:
        try:
            time_str = datetime.strptime(ts, "%Y%m%d_%H%M%S").strftime("%Y-%m-%d %H:%M:%S")
        except ValueError:
            time_str = ""
    return {"undo_id": undo_id, "action": action, "time": time_str}


def list_recent(data_dir: str, limit: int = 15) -> list:
    """最近可撤回操作 (新→旧): [{time, action, subcat, n}]。"""
    out = []
    # Excel 撤回日志
    for p, name in _iter_snap_files(data_dir):
        info = _snap_info_from_filename(name)
        rows = _read_snap(p)
        subcats = list(dict.fromkeys("未分类" if r["subcat"] == UNCAT_KEY else r["subcat"] for r in rows))
        out.append({
            "time": info.get("time", ""),
            "action": info.get("action", ""),
            "subcat": "、".join(subcats),
            "n": len(rows),
            "undo_id": info.get("undo_id", ""),
        })
    # 旧 jsonl (兼容)
    for e in reversed(load_all(data_dir)):
        snaps = e.get("snapshots") or []
        if not snaps:
            continue
        subcats = ["未分类" if s.get("subcat") == UNCAT_KEY else s.get("subcat", "") for s in snaps]
        out.append({
            "time": e.get("time", ""),
            "action": e.get("action", ""),
            "subcat": "、".join(dict.fromkeys(subcats)),
            "n": len(snaps),
        })
    out.sort(key=lambda x: x.get("time", ""), reverse=True)
    return out[:limit]


# ── 旧 jsonl 兼容 ─────────────────────────────────────────

def load_all(data_dir: str) -> list:
    p = _path(data_dir)
    if not os.path.exists(p):
        return []
    entries = []
    with open(p, encoding="utf-8") as f:
        for line in f:
            line = line.strip()
            if not line:
                continue
            try:
                entries.append(json.loads(line))
            except json.JSONDecodeError:
                continue
    return entries


def _write_all(data_dir: str, entries: list):
    p = _path(data_dir)
    os.makedirs(os.path.dirname(p), exist_ok=True)
    with open(p, "w", encoding="utf-8") as f:
        for e in entries:
            f.write(json.dumps(e, ensure_ascii=False) + "\n")


# ── 撤回执行 ──────────────────────────────────────────────

def _save_rows(data_dir: str, subcat: str, headers: list, current: list):
    from warehouse.excel_store import ExcelStore
    store = ExcelStore(data_dir)
    if subcat == UNCAT_KEY:
        from warehouse import unclassified
        unclassified._save(data_dir, headers, current)
    else:
        store.save(subcat, headers, current)


def _undo_from_excel(data_dir: str, path: str, undo_id: str, item_rows: list | None = None) -> dict:
    """按撤回日志 Excel 撤回整笔或指定明细行。"""
    snap_rows = _read_snap(path)
    requested = {(str(x.get("subcat", "")), int(x.get("row", -1))) for x in (item_rows or [])}
    partial = bool(requested)
    headers = [label for _, label in COMMON_FIELDS]
    restored = []
    skipped = []
    remaining = []
    # 按子分类分组
    by_subcat: dict[str, list] = {}
    for r in snap_rows:
        by_subcat.setdefault(r["subcat"], []).append(r)
    for subcat, group in by_subcat.items():
        if partial:
            targets = [r for r in group if (r["subcat"], r["orig_row"]) in requested]
        else:
            targets = group
        if not targets:
            remaining.extend(group)
            continue
        # 恢复目标行到当前表
        from warehouse.excel_store import ExcelStore
        store = ExcelStore(data_dir)
        if subcat == UNCAT_KEY:
            from warehouse import unclassified
            cur_headers, current = unclassified.load(data_dir)
        else:
            _, current = store.load(subcat)
        ok_rows = []
        for r in targets:
            if 0 <= r["orig_row"] < len(current):
                current[r["orig_row"]] = r["fields"]
                ok_rows.append(r)
            else:
                skipped.append(r)
        if subcat == UNCAT_KEY:
            unclassified._save(data_dir, cur_headers, current)
        else:
            store.save(subcat, headers, current)
        if ok_rows:
            restored.append({"subcat": subcat, "rows": sorted(r["orig_row"] for r in ok_rows)})
        # 未请求的行保留 (继续单项撤回)；写回失败的行也保留
        remaining.extend(r for r in group if r not in ok_rows)
    # 剩余行重写文件；全部恢复则删除文件
    if remaining:
        _write_snap(path, remaining)
        status = "部分撤回" if partial or skipped else "已撤回"
    else:
        try:
            os.remove(path)
        except OSError:
            pass
        status = "已撤回"
    entry = {"undo_id": undo_id, "status": status, "restored": restored}
    if skipped:
        entry["message"] = f"{len(skipped)} 行因表格结构变化未能恢复，已保留在撤回日志中"
    return entry


def undo_by_id(data_dir: str, undo_id: str, item_rows: list | None = None) -> dict | None:
    """按唯一撤回 ID 撤回整笔或指定明细行。item_rows 为 [{subcat,row}]。"""
    p = _snap_file_by_undo(data_dir, undo_id)
    if p:
        return _undo_from_excel(data_dir, p, undo_id, item_rows)
    # 旧 jsonl 兼容路径
    entries = load_all(data_dir)
    idx = next((i for i, e in enumerate(entries) if e.get("undo_id") == undo_id), -1)
    if idx < 0:
        return None
    entry = entries[idx]
    requested = {(str(x.get("subcat", "")), int(x.get("row", -1))) for x in (item_rows or [])}
    partial = bool(requested)
    from warehouse.excel_store import ExcelStore
    store = ExcelStore(data_dir)
    headers = [label for _, label in COMMON_FIELDS]
    restored = []
    remaining = []
    for snap in entry.get("snapshots") or []:
        subcat = snap.get("subcat", "")
        old_rows = snap.get("old_rows") or []
        if not partial:
            targets = set(range(len(old_rows)))
        else:
            already = set(snap.get("undone_rows") or [])
            targets = {row for sc, row in requested if sc == subcat and row not in already}
        if not targets:
            remaining.append(snap)
            continue
        if subcat == UNCAT_KEY:
            from warehouse import unclassified
            current_headers, current = unclassified.load(data_dir)
            for row in targets:
                if 0 <= row < len(current) and row < len(old_rows):
                    current[row] = old_rows[row]
            unclassified._save(data_dir, current_headers, current)
        else:
            _, current = store.load(subcat)
            for row in targets:
                if 0 <= row < len(current) and row < len(old_rows):
                    current[row] = old_rows[row]
            store.save(subcat, headers, current)
        restored.append({"subcat": subcat, "rows": sorted(targets)})
        if partial:
            all_undone = set(snap.get("undone_rows") or []) | targets
            if len(all_undone) < len(old_rows):
                remaining.append({**snap, "undone_rows": sorted(all_undone)})
    if remaining:
        entries[idx] = {**entry, "snapshots": remaining, "status": "部分撤回"}
    else:
        entries.pop(idx)
    _write_all(data_dir, entries)
    entry["restored"] = restored
    entry["status"] = "部分撤回" if remaining else "已撤回"
    return entry


def undo(data_dir: str, time: str) -> dict | None:
    """旧版按时间撤回 (顶部"撤回"弹窗)。Excel 文件按时间戳匹配。"""
    for p, name in _iter_snap_files(data_dir):
        info = _snap_info_from_filename(name)
        if info.get("time") == time:
            entry = _undo_from_excel(data_dir, p, info.get("undo_id", ""), None)
            entry["time"] = time
            entry["action"] = info.get("action", "")
            return entry
    # 旧 jsonl 兼容路径
    entries = load_all(data_dir)
    idx = -1
    for i in range(len(entries) - 1, -1, -1):
        if entries[i].get("time") == time:
            idx = i
            break
    if idx < 0:
        return None
    entry = entries.pop(idx)
    _write_all(data_dir, entries)
    from warehouse.excel_store import ExcelStore
    store = ExcelStore(data_dir)
    headers = [label for _, label in COMMON_FIELDS]
    restored = []
    for snap in entry.get("snapshots") or []:
        subcat = snap.get("subcat", "")
        old_rows = snap.get("old_rows") or []
        if subcat == UNCAT_KEY:
            from warehouse import unclassified
            unclassified._save(data_dir, unclassified.HEADERS, old_rows)
        else:
            store.save(subcat, headers, old_rows)
        restored.append(subcat)
    entry["restored"] = restored
    return entry
