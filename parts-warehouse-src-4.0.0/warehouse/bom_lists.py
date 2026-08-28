# -*- coding: utf-8 -*-
"""BOM清单 Excel 持久化。

每个已确认的 BOM 清单独立保存为 data/BOM清单/<名称>.xlsx。
Excel 既供用户查看，也保存仓库匹配绑定所需的 JSON 原始字段。
"""
from __future__ import annotations

import json
import os
import re
import uuid
from datetime import datetime

from openpyxl import Workbook, load_workbook

DIR_NAME = "BOM清单"
LEGACY_NAME = "bom_lists.json"
SHEET_NAME = "BOM清单"
HEADERS = [
    "序号", "需求名称/型号", "品牌", "封装", "单板用量", "规格参数",
    "器件分类", "仓库子分类", "仓库行号", "仓库型号", "仓库品牌",
    "仓库封装", "当前库存", "仓库规格", "匹配状态", "来源图片/行", "item_json", "selected_json",
]
INVALID_FILENAME_RE = re.compile(r'[<>:"/\\|?*\x00-\x1f]')


def _dir(data_dir: str) -> str:
    return os.path.join(data_dir, DIR_NAME)


def _legacy_path(data_dir: str) -> str:
    return os.path.join(data_dir, LEGACY_NAME)


def _clean_filename(value: str) -> str:
    value = INVALID_FILENAME_RE.sub("_", str(value or "").strip())
    value = value.rstrip(" .")[:100]
    return value or "未命名BOM"


def _json_cell(value) -> str:
    return json.dumps(value or {}, ensure_ascii=False, separators=(",", ":"))


def _json_value(value, default):
    if not isinstance(value, str):
        return value if isinstance(value, type(default)) else default
    try:
        parsed = json.loads(value)
        return parsed if isinstance(parsed, type(default)) else default
    except (TypeError, ValueError):
        return default


def _meta(ws) -> dict:
    return {
        "id": str(ws["B4"].value or ""),
        "name": str(ws["B1"].value or ""),
        "source": str(ws["B2"].value or ""),
        "created_at": str(ws["B3"].value or ""),
        "updated_at": str(ws["B5"].value or ""),
    }


def _path_for(data_dir: str, item: dict) -> str:
    return os.path.join(_dir(data_dir), _clean_filename(item.get("name")) + ".xlsx")


def _read_xlsx(path: str) -> dict | None:
    try:
        wb = load_workbook(path, read_only=True, data_only=True)
        ws = wb[SHEET_NAME] if SHEET_NAME in wb.sheetnames else wb.active
        meta = _meta(ws)
        if not meta["id"] or not meta["name"]:
            return None
        rows = []
        for values in ws.iter_rows(min_row=7, values_only=True):
            if not any(v not in (None, "") for v in values):
                continue
            # 新格式有“匹配状态”列：JSON 在 Q/R；旧格式 JSON 在 P/Q。
            new_format = len(values) > 17
            item_cell = values[16] if new_format else (values[15] if len(values) > 15 else "")
            selected_cell = values[17] if new_format else (values[16] if len(values) > 16 else "")
            item = _json_value(item_cell, {})
            selected = _json_value(selected_cell, {})
            has_item_json = isinstance(item_cell, str) and item_cell.strip().startswith("{")
            has_selected_json = isinstance(selected_cell, str) and selected_cell.strip().startswith("{")
            if not item:
                item = {
                    "name": values[1] or "", "brand": values[2] or "",
                    "package": values[3] or "", "qty": values[4] or "",
                    "spec": values[5] or "", "cat_key": values[6] or "",
                }
            if not selected and not has_selected_json and values[7]:
                selected = {
                    "subcat": values[7] or "", "row": values[8] or -1,
                    "snapshot": {"name": values[9] or "", "brand": values[10] or "",
                                 "package": values[11] or "", "spec": values[13] or ""},
                }
            rows.append({"item": item, "selected": selected})
        wb.close()
        meta["items"] = rows
        return meta
    except (OSError, ValueError, KeyError):
        return None


def _write_xlsx(data_dir: str, item: dict, path: str | None = None) -> dict:
    os.makedirs(_dir(data_dir), exist_ok=True)
    path = path or _path_for(data_dir, item)
    wb = Workbook()
    ws = wb.active
    ws.title = SHEET_NAME
    ws["A1"], ws["B1"] = "BOM清单名称", item["name"]
    ws["A2"], ws["B2"] = "来源文件", item.get("source", "")
    ws["A3"], ws["B3"] = "创建时间", item["created_at"]
    ws["A4"], ws["B4"] = "清单ID", item["id"]
    ws["A5"], ws["B5"] = "更新时间", item["updated_at"]
    for col, title in enumerate(HEADERS, 1):
        ws.cell(6, col, title)
    for index, entry in enumerate(item.get("items", []), 1):
        req = entry.get("item") or {}
        selected = entry.get("selected") or {}
        snapshot = selected.get("snapshot") or {}
        values = [
            index, req.get("name", ""), req.get("brand", ""), req.get("package", ""),
            req.get("qty", ""), req.get("spec", ""), req.get("cat_key", ""),
            selected.get("subcat", ""), selected.get("row", ""), snapshot.get("name", ""),
            snapshot.get("brand", ""), snapshot.get("package", ""), snapshot.get("qty", ""), snapshot.get("spec", ""),
            "已匹配" if selected.get("subcat") else "未匹配/待补料",
            req.get("source_image", req.get("raw", "")), _json_cell(req), _json_cell(selected),
        ]
        for col, value in enumerate(values, 1):
            ws.cell(6 + index, col, value)
    ws.freeze_panes = "A7"
    ws.auto_filter.ref = f"A6:R{max(6, 6 + len(item.get('items', [])))}"
    widths = [8, 24, 16, 14, 12, 32, 18, 28, 12, 24, 16, 14, 12, 32, 18, 24, 28, 28]
    for i, width in enumerate(widths, 1):
        ws.column_dimensions[chr(64 + i) if i <= 26 else "R"].width = width
    tmp = path + ".tmp"
    wb.save(tmp)
    os.replace(tmp, path)
    return item


def _legacy_read(data_dir: str) -> list[dict]:
    path = _legacy_path(data_dir)
    if not os.path.exists(path):
        return []
    try:
        with open(path, "r", encoding="utf-8") as f:
            data = json.load(f)
        return data if isinstance(data, list) else []
    except (OSError, ValueError):
        return []


def _migrate_legacy(data_dir: str) -> list[dict]:
    legacy = _legacy_read(data_dir)
    if not legacy:
        return []
    os.makedirs(_dir(data_dir), exist_ok=True)
    migrated = []
    for item in legacy:
        if not item.get("id"):
            item["id"] = uuid.uuid4().hex
        item.setdefault("created_at", datetime.now().isoformat(timespec="seconds"))
        item.setdefault("updated_at", item["created_at"])
        _write_xlsx(data_dir, item)
        migrated.append(item)
    try:
        os.replace(_legacy_path(data_dir), _legacy_path(data_dir) + ".migrated")
    except OSError:
        pass
    return migrated


def _all_xlsx(data_dir: str) -> list[dict]:
    folder = _dir(data_dir)
    if not os.path.isdir(folder):
        return []
    result = []
    for name in os.listdir(folder):
        if not name.lower().endswith(".xlsx"):
            continue
        item = _read_xlsx(os.path.join(folder, name))
        if item:
            item["_path"] = os.path.join(folder, name)
            result.append(item)
    result.sort(key=lambda x: x.get("updated_at", ""), reverse=True)
    return result


def list_lists(data_dir: str) -> list[dict]:
    _migrate_legacy(data_dir)
    return [{
        "id": x.get("id", ""), "name": x.get("name", ""), "source": x.get("source", ""),
        "created_at": x.get("created_at", ""), "updated_at": x.get("updated_at", ""),
        "count": len(x.get("items", [])),
    } for x in _all_xlsx(data_dir)]


def get_list(data_dir: str, list_id: str) -> dict | None:
    _migrate_legacy(data_dir)
    return next((x for x in _all_xlsx(data_dir) if x.get("id") == list_id), None)


def save_list(data_dir: str, payload: dict) -> dict:
    name = str(payload.get("name") or "").strip()[:120]
    rows = payload.get("items") or []
    if not name:
        raise ValueError("BOM 清单名称不能为空")
    if not isinstance(rows, list) or not rows:
        raise ValueError("BOM 清单至少需要一条物料")
    now = datetime.now().isoformat(timespec="seconds")
    existing = get_list(data_dir, str(payload.get("id") or "")) if payload.get("id") else None
    item = {
        "id": str(payload.get("id") or uuid.uuid4().hex), "name": name,
        "source": str(payload.get("source") or "").strip()[:200],
        "created_at": (existing or {}).get("created_at", now), "updated_at": now,
        "items": rows,
    }
    target = _path_for(data_dir, item)
    if existing and existing.get("_path") and existing["_path"] != target:
        try:
            os.remove(existing["_path"])
        except OSError:
            pass
    return _write_xlsx(data_dir, item, target)


def delete_list(data_dir: str, list_id: str) -> bool:
    item = get_list(data_dir, list_id)
    if not item:
        return False
    try:
        os.remove(item["_path"])
    except OSError:
        return False
    return True
